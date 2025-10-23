from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Avg, Q, F
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime, timedelta
from decimal import Decimal
from .models import *
from django.db import models
import json
import os

# ==================== HELPER FUNCTIONS ====================

def get_dashboard_by_role(user):
    """Retourne l'URL du dashboard approprié selon le rôle de l'utilisateur"""
    dashboard_map = {
        'DG': 'dashboard_dg',
        'DAF': 'dashboard_daf',
        'RH': 'dashboard_rh',
        'STOCK': 'dashboard_stock',
        'CAISSIER': 'dashboard_caisse',
        'MARKETING': 'dashboard_marketing',
    }
    return dashboard_map.get(user.role, 'home')

# Page d'accueil
def home(request):
    return render(request, 'home.html')

# Page À propos
def about(request):
    return render(request, 'about.html')

# Page Contact
def contact(request):
    return render(request, 'contact.html')

# Page de connexion
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            user.derniere_connexion_custom = timezone.now()
            user.save()
            
            # 🕐 POINTAGE AUTOMATIQUE D'ARRIVÉE (NOUVELLE SESSION)
            today = timezone.now().date()
            current_time = timezone.now().time()
            
            # Vérifier si l'employé est dans les heures de travail
            if user.heure_debut_travail <= current_time <= user.heure_fin_travail:
                # Importer le modèle SessionPresence
                from .models import SessionPresence
                
                # Créer une nouvelle session de connexion
                SessionPresence.objects.create(
                    employe=user,
                    date=today,
                    heure_connexion=current_time
                )
                
                # Créer ou récupérer la présence du jour
                presence, created = Presence.objects.get_or_create(
                    employe=user,
                    date=today,
                    defaults={
                        'heure_premiere_arrivee': current_time,
                        'tolerance_retard': 60  # 1 heure par défaut
                    }
                )
                
                # Si c'est la première connexion du jour, enregistrer l'heure
                if not created and not presence.heure_premiere_arrivee:
                    presence.heure_premiere_arrivee = current_time
                    presence.save()
            
            # Redirection selon le rôle
            if user.role == 'DG':
                return redirect('dashboard_dg')
            elif user.role == 'DAF':
                return redirect('dashboard_daf')
            elif user.role == 'RH':
                return redirect('dashboard_rh')
            elif user.role == 'STOCK':
                return redirect('dashboard_stock')
            elif user.role == 'CAISSIER':
                return redirect('dashboard_caisse')
            elif user.role == 'MARKETING':
                return redirect('dashboard_marketing')
            elif user.role == 'ANALYSTE':
                return redirect('dashboard_analytics')
            else:
                return redirect('dashboard')
        else:
            messages.error(request, 'Identifiant ou mot de passe incorrect')
    
    return render(request, 'login.html')

# Déconnexion
def logout_view(request):
    # 🕔 POINTAGE AUTOMATIQUE DE DÉPART (FERMER LA SESSION EN COURS)
    if request.user.is_authenticated:
        from .models import SessionPresence
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        try:
            # Trouver la dernière session active (sans heure de déconnexion)
            session_active = SessionPresence.objects.filter(
                employe=request.user,
                date=today,
                heure_deconnexion__isnull=True
            ).last()
            
            if session_active:
                # Fermer la session en enregistrant l'heure de déconnexion
                session_active.heure_deconnexion = current_time
                session_active.save()  # Calcule automatiquement la durée
                
                # Mettre à jour la présence du jour
                try:
                    presence = Presence.objects.get(employe=request.user, date=today)
                    presence.heure_derniere_depart = current_time
                    presence.save()  # Recalcule le statut automatiquement
                except Presence.DoesNotExist:
                    pass
                    
        except Exception as e:
            # En cas d'erreur, continuer la déconnexion normalement
            pass
    
    logout(request)
    return redirect('home')

# Tableau de bord principal (sélection de profil)
@login_required
def dashboard(request):
    context = {
        'user': request.user,
    }
    return render(request, 'dashboard/main.html', context)

# Tableau de bord Direction Générale
@login_required
def dashboard_dg(request):
    # Vérifier les permissions
    if request.user.role != 'DG':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Direction Générale.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Ventes du jour
    ventes_jour = Vente.objects.filter(date_vente__date=today)
    ca_jour = ventes_jour.aggregate(total=Sum('montant_final'))['total'] or 0
    
    # Ventes du mois
    ventes_mois = Vente.objects.filter(
        date_vente__month=current_month,
        date_vente__year=current_year
    )
    ca_mois = ventes_mois.aggregate(total=Sum('montant_final'))['total'] or 0
    nb_transactions = ventes_mois.count()
    
    # Panier moyen
    panier_moyen = ca_mois / nb_transactions if nb_transactions > 0 else 0
    
    # ROI (Return on Investment)
    total_ventes = ventes_mois.aggregate(total=Sum('montant_final'))['total'] or 0
    total_achats = 0
    for vente in ventes_mois:
        for ligne in vente.lignes.all():
            total_achats += ligne.produit.prix_achat * ligne.quantite
    
    roi = ((total_ventes - total_achats) / total_achats * 100) if total_achats > 0 else 0
    
    # Évolution du chiffre d'affaires (6 derniers mois)
    ca_evolution = []
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        ca = Vente.objects.filter(
            date_vente__month=date.month,
            date_vente__year=date.year
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        ca_evolution.append({
            'mois': date.strftime('%b'),
            'ca': float(ca)
        })
    
    # Top produits par revenus
    top_produits = []
    produits_vendus = LigneVente.objects.filter(
        vente__date_vente__month=current_month
    ).values('produit__nom', 'produit__categorie').annotate(
        revenus=Sum('montant_ligne')
    ).order_by('-revenus')[:4]
    
    for item in produits_vendus:
        top_produits.append({
            'nom': item['produit__nom'],
            'categorie': item['produit__categorie'],
            'revenus': float(item['revenus'])
        })
    
    # Analyse des marges
    marges_data = []
    mois_labels = []
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        ventes = Vente.objects.filter(
            date_vente__month=date.month,
            date_vente__year=date.year
        )
        
        revenus = 0
        couts = 0
        benefices = 0
        
        for vente in ventes:
            revenus += float(vente.montant_final)
            for ligne in vente.lignes.all():
                couts += float(ligne.produit.prix_achat * ligne.quantite)
        
        benefices = revenus - couts
        
        mois_labels.append(date.strftime('%b'))
        marges_data.append({
            'revenus': revenus,
            'couts': couts,
            'benefices': benefices
        })
    
    # Indicateurs opérationnels RÉELS
    # Taux de rotation des stocks (nombre de ventes / stock moyen)
    nb_ventes_mois = Vente.objects.filter(date_vente__month=current_month).count()
    stock_moyen = Produit.objects.filter(stock_actuel__gt=0).aggregate(avg=Avg('stock_actuel'))['avg'] or 1
    taux_rotation_stocks = round(nb_ventes_mois / float(stock_moyen), 2) if stock_moyen > 0 else 0
    
    # Temps moyen de traitement à la caisse (basé sur le nombre de lignes de vente)
    lignes_moyennes = LigneVente.objects.filter(vente__date_vente__month=current_month).count() / max(nb_ventes_mois, 1)
    temps_moyen_caisse = round(lignes_moyennes * 0.5, 1)  # 0.5 min par article
    
    # Satisfaction client (basé sur les réclamations)
    total_clients = Client.objects.count() or 1
    reclamations = Reclamation.objects.filter(date_creation__month=current_month).count()
    satisfaction_client = round(100 - (reclamations / total_clients * 100), 1)
    
    # Productivité employés (ventes par employé actif)
    employes_actifs = Employe.objects.filter(est_actif=True).count() or 1
    productivite_employes = round((nb_ventes_mois / employes_actifs), 0)
    
    # Taux de déchets/pertes (produits en rupture ou critique)
    total_produits = Produit.objects.count() or 1
    produits_critiques = Produit.objects.filter(stock_actuel__lt=10).count()
    taux_dechet = round((produits_critiques / total_produits * 100), 1)
    
    context = {
        'ca_jour': ca_jour,
        'ca_mois': ca_mois,
        'benefice_mois': ca_mois * Decimal('0.182'),
        'nb_transactions': nb_transactions,
        'panier_moyen': panier_moyen,
        'roi': roi,
        'ca_evolution': ca_evolution,
        'top_produits': top_produits,
        'mois_labels': mois_labels,
        'marges_data': marges_data,
        'taux_rotation_stocks': taux_rotation_stocks,
        'temps_moyen_caisse': temps_moyen_caisse,
        'satisfaction_client': satisfaction_client,
        'productivite_employes': productivite_employes,
        'taux_dechet': taux_dechet,
    }
    
    return render(request, 'dashboard/dg.html', context)

# Tableau de bord DAF (Directeur Administratif et Financier)
@login_required
def dashboard_daf(request):
    # Vérifier les permissions
    if request.user.role != 'DAF':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Financier.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Chiffre d'affaires mensuel (RÉEL)
    ca_mois = Vente.objects.filter(
        date_vente__month=current_month,
        date_vente__year=current_year
    ).aggregate(total=Sum('montant_final'))['total'] or 0
    
    # Calculer la marge bénéficiaire RÉELLE (revenus - coûts)
    revenus_mois = float(ca_mois)
    couts_mois = 0
    for vente in Vente.objects.filter(date_vente__month=current_month):
        for ligne in vente.lignes.all():
            couts_mois += float(ligne.produit.prix_achat * ligne.quantite)
    
    marge_beneficiaire = ((revenus_mois - couts_mois) / revenus_mois * 100) if revenus_mois > 0 else 0
    
    # Charges mensuelles RÉELLES
    charges_mensuelles = couts_mois
    
    # Trésorerie RÉELLE (CA total - Charges)
    ca_total = Vente.objects.aggregate(total=Sum('montant_final'))['total'] or 0
    tresorerie = float(ca_total) - charges_mensuelles
    
    # Évolution du CA (6 derniers mois)
    ca_evolution = []
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        ca = Vente.objects.filter(
            date_vente__month=date.month,
            date_vente__year=date.year
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        ca_evolution.append({
            'mois': date.strftime('%b'),
            'ca': float(ca),
            'objectif': float(ca) * 0.95
        })
    
    # Analyse des marges (brute et nette)
    marges_data = []
    for i in range(5, -1, -1):
        date = today - timedelta(days=i*30)
        ca = Vente.objects.filter(
            date_vente__month=date.month,
            date_vente__year=date.year
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        
        marges_data.append({
            'mois': date.strftime('%b'),
            'marge_brute': float(ca) * 0.28,
            'marge_nette': float(ca) * 0.18
        })
    
    # Modes de paiement RÉELS
    moyens_paiement = Vente.objects.filter(
        date_vente__month=current_month
    ).values('moyen_paiement').annotate(
        total=Count('id')
    )
    
    paiement_data = []
    total_ventes = Vente.objects.filter(date_vente__month=current_month).count()
    
    for mp in moyens_paiement:
        pourcentage = (mp['total'] / total_ventes * 100) if total_ventes > 0 else 0
        paiement_data.append({
            'moyen': mp['moyen_paiement'],
            'pourcentage': round(pourcentage, 1)
        })
    
    context = {
        'ca_mois': ca_mois,
        'marge_beneficiaire': round(marge_beneficiaire, 2),
        'charges_mensuelles': charges_mensuelles,
        'tresorerie': tresorerie,
        'ca_evolution': ca_evolution,
        'marges_data': marges_data,
        'paiement_data': paiement_data,
    }
    
    return render(request, 'dashboard/daf.html', context)

# Tableau de bord RH
@login_required
def dashboard_rh(request):
    # Vérifier que l'utilisateur a accès au dashboard RH
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module RH.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    
    # Statistiques employés (VRAIES DONNÉES uniquement)
    total_employes = Employe.objects.filter(est_actif=True).count()
    presences_jour = Presence.objects.filter(date=today, statut__in=['PRESENT', 'RETARD']).count()
    taux_presence = (presences_jour / total_employes * 100) if total_employes > 0 else 0
    
    # Congés en cours (RÉELS)
    conges_en_cours = Conge.objects.filter(
        date_debut__lte=today,
        date_fin__gte=today,
        statut='APPROUVE'
    ).count()
    
    # Formations actives (RÉELLES)
    formations_actives = Formation.objects.filter(
        date_debut__lte=today,
        date_fin__gte=today,
        est_terminee=False
    ).count()
    
    # Liste des employés RÉELS
    employes = Employe.objects.filter(est_actif=True).order_by('-date_joined')
    
    # Activités récentes RÉELLES (derniers 7 jours)
    activites = []
    
    # Nouveaux employés (derniers 7 jours)
    nouveaux = Employe.objects.filter(
        date_embauche__gte=today - timedelta(days=7)
    ).order_by('-date_embauche')[:3]
    
    for emp in nouveaux:
        temps = (today - emp.date_embauche).days
        activites.append({
            'type': 'employe',
            'titre': 'Nouvel employé ajouté',
            'details': f"{emp.get_full_name()} - {emp.get_role_display()}",
            'temps': f"Il y a {temps} jour{'s' if temps > 1 else ''}"
        })
    
    # Congés récents
    conges_recents = Conge.objects.filter(statut='APPROUVE').order_by('-date_demande')[:2]
    for conge in conges_recents:
        duree = (conge.date_fin - conge.date_debut).days
        activites.append({
            'type': 'conge',
            'titre': 'Congé approuvé',
            'details': f"{conge.employe.get_full_name()} - {duree} jours",
            'temps': "Il y a 4h"
        })
    
    # Formations récentes
    formations_recentes = Formation.objects.filter(est_terminee=True).order_by('-date_fin')[:1]
    for formation in formations_recentes:
        activites.append({
            'type': 'formation',
            'titre': 'Formation terminée',
            'details': f"{formation.titre} - {formation.nombre_participants} participants",
            'temps': "Ici"
        })
    
    context = {
        'total_employes': total_employes,
        'presences_jour': presences_jour,
        'taux_presence': taux_presence,
        'conges_en_cours': conges_en_cours,
        'formations_actives': formations_actives,
        'employes': employes,
        'activites': activites,
    }
    
    return render(request, 'dashboard/rh.html', context)

# Tableau de bord Stock
@login_required
def dashboard_stock(request):
    """Dashboard Stock enrichi avec KPIs avancés - Sprint 1 Jour 11-12"""
    # Vérifier les permissions
    if request.user.role != 'STOCK':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Stock.")
        return redirect('dashboard')
    
    # ========== STATISTIQUES DE BASE ==========
    total_produits = Produit.objects.filter(est_actif=True).count()
    produits_rupture = Produit.objects.filter(stock_actuel=0, est_actif=True).count()
    produits_critiques = Produit.objects.filter(stock_actuel__gt=0, stock_actuel__lte=10, est_actif=True).count()
    
    # ========== VALEUR DU STOCK ==========
    valeur_stock_achat = Decimal('0')
    valeur_stock_vente = Decimal('0')
    for produit in Produit.objects.filter(est_actif=True):
        valeur_stock_achat += produit.valeur_stock()
        valeur_stock_vente += produit.valeur_stock_vente()
    
    marge_potentielle = valeur_stock_vente - valeur_stock_achat
    taux_marge = (marge_potentielle / valeur_stock_achat * 100) if valeur_stock_achat > 0 else 0
    
    # ========== COMMANDES FOURNISSEURS ==========
    commandes_en_attente = CommandeFournisseur.objects.filter(statut='EN_ATTENTE').count()
    commandes_validees = CommandeFournisseur.objects.filter(statut='VALIDEE').count()
    commandes_en_cours = commandes_en_attente + commandes_validees
    
    # ========== ALERTES ACTIVES ==========
    alertes_actives = AlerteStock.objects.filter(est_resolue=False).count()
    alertes_rupture = AlerteStock.objects.filter(type_alerte='RUPTURE', est_resolue=False).count()
    alertes_seuil = AlerteStock.objects.filter(type_alerte='SEUIL_CRITIQUE', est_resolue=False).count()
    
    # ========== TOP PRODUITS À RÉAPPROVISIONNER ==========
    produits_a_reappro = []
    for produit in Produit.objects.filter(est_actif=True):
        if produit.besoin_reapprovisionnement():
            produits_a_reappro.append({
                'id': produit.id,
                'nom': produit.nom,
                'stock_actuel': produit.stock_actuel,
                'seuil': produit.seuil_reapprovisionnement,
                'quantite_recommandee': produit.quantite_a_commander(),
                'fournisseur': produit.fournisseur_principal.nom if produit.fournisseur_principal else 'Non défini',
                'urgence': 'CRITIQUE' if produit.stock_actuel <= produit.stock_minimum else 'MOYEN'
            })
    produits_a_reappro = sorted(produits_a_reappro, key=lambda x: x['stock_actuel'])[:10]
    
    # ========== MOUVEMENTS RÉCENTS ==========
    mouvements_recents = MouvementStock.objects.all().select_related('produit', 'employe').order_by('-date_mouvement')[:10]
    
    # ========== STATISTIQUES PAR CATÉGORIE ==========
    categories_stats = []
    # Utiliser les catégories définies dans le modèle Produit.CATEGORIES
    for cat_code, cat_label in Produit.CATEGORIES:
        produits_cat = Produit.objects.filter(categorie=cat_code, est_actif=True)
        nb_produits = produits_cat.count()
        if nb_produits > 0:
            valeur_cat = sum(p.valeur_stock() for p in produits_cat)
            categories_stats.append({
                'nom': cat_label,
                'code': cat_code,
                'nb_produits': nb_produits,
                'valeur': valeur_cat,
                'pourcentage': (valeur_cat / valeur_stock_achat * 100) if valeur_stock_achat > 0 else 0
            })
    categories_stats = sorted(categories_stats, key=lambda x: x['valeur'], reverse=True)[:5]
    
    # ========== TAUX DE ROTATION STOCK (7 derniers jours) ==========
    from datetime import timedelta
    today = timezone.now().date()
    date_debut = today - timedelta(days=7)
    
    sorties_semaine = MouvementStock.objects.filter(
        type_mouvement='SORTIE',
        date_mouvement__date__gte=date_debut
    ).aggregate(total=Sum('quantite'))['total'] or 0
    
    stock_moyen = Produit.objects.filter(est_actif=True).aggregate(total=Sum('stock_actuel'))['total'] or 1
    taux_rotation = (sorties_semaine / stock_moyen * 100) if stock_moyen > 0 else 0
    
    # ========== PERFORMANCE FOURNISSEURS ==========
    fournisseurs_stats = []
    fournisseurs = Fournisseur.objects.filter(est_actif=True)[:5]
    for f in fournisseurs:
        commandes = CommandeFournisseur.objects.filter(fournisseur=f)
        nb_commandes = commandes.count()
        if nb_commandes > 0:
            livrees_a_temps = commandes.filter(
                statut='LIVREE',
                date_livraison_reelle__lte=F('date_livraison_prevue')
            ).count()
            taux_ponctualite = (livrees_a_temps / nb_commandes * 100) if nb_commandes > 0 else 0
            fournisseurs_stats.append({
                'nom': f.nom,
                'nb_commandes': nb_commandes,
                'taux_ponctualite': round(taux_ponctualite, 1),
                'montant_total': f.montant_total_commandes()
            })
    
    # Liste des produits pour tableau
    produits = Produit.objects.filter(est_actif=True).select_related('fournisseur_principal').order_by('nom')[:50]
    
    context = {
        'total_produits': total_produits,
        'produits_rupture': produits_rupture,
        'produits_critiques_count': produits_critiques,
        'valeur_stock_achat': valeur_stock_achat,
        'valeur_stock_vente': valeur_stock_vente,
        'marge_potentielle': marge_potentielle,
        'taux_marge': round(taux_marge, 1),
        'commandes_en_cours': commandes_en_cours,
        'commandes_en_attente': commandes_en_attente,
        'commandes_validees': commandes_validees,
        'alertes_actives': alertes_actives,
        'alertes_rupture': alertes_rupture,
        'alertes_seuil': alertes_seuil,
        'produits_a_reappro': produits_a_reappro,
        'mouvements_recents': mouvements_recents,
        'categories_stats': categories_stats,
        'taux_rotation': round(taux_rotation, 1),
        'fournisseurs_stats': fournisseurs_stats,
        'produits': produits,
    }
    
    return render(request, 'dashboard/stock.html', context)

# Tableau de bord Caisse
@login_required
def dashboard_caisse(request):
    # Vérifier les permissions
    if request.user.role != 'CAISSIER':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Caisse.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    
    # Ventes du jour RÉELLES
    ventes_jour = Vente.objects.filter(date_vente__date=today)
    ca_jour = ventes_jour.aggregate(total=Sum('montant_final'))['total'] or 0
    nb_transactions = ventes_jour.count()
    panier_moyen = ca_jour / nb_transactions if nb_transactions > 0 else 0
    
    # Caisses actives RÉELLES (nombre de caissiers actifs aujourd'hui)
    caisses_actives = Employe.objects.filter(
        role='CAISSIER',
        est_actif=True
    ).count()
    total_caisses = Employe.objects.filter(role='CAISSIER').count()
    
    # Transactions récentes
    transactions_recentes = Vente.objects.filter(date_vente__date=today).order_by('-date_vente')[:5]
    
    # Moyens de paiement
    moyens_paiement = Vente.objects.filter(date_vente__date=today).values(
        'moyen_paiement'
    ).annotate(
        total=Count('id')
    )
    
    paiement_stats = []
    for mp in moyens_paiement:
        pourcentage = (mp['total'] / nb_transactions * 100) if nb_transactions > 0 else 0
        paiement_stats.append({
            'moyen': mp['moyen_paiement'],
            'pourcentage': round(pourcentage, 0)
        })
    
    context = {
        'ca_jour': ca_jour,
        'nb_transactions': nb_transactions,
        'panier_moyen': panier_moyen,
        'caisses_actives': caisses_actives,
        'total_caisses': total_caisses,
        'transactions_recentes': transactions_recentes,
        'paiement_stats': paiement_stats,
    }
    
    return render(request, 'dashboard/caisse.html', context)

# Tableau de bord Marketing (Fidélisation)
@login_required
def dashboard_marketing(request):
    # Vérifier les permissions
    if request.user.role != 'MARKETING':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Marketing.")
        return redirect('dashboard')
    
    # Statistiques clients RÉELLES
    clients_fideles = Client.objects.count()
    points_distribues = Client.objects.aggregate(total=Sum('points_fidelite'))['total'] or 0
    promotions_actives = Promotion.objects.filter(est_active=True).count()
    reclamations = Reclamation.objects.filter(statut='EN_COURS').count()
    
    # Nouveaux clients ce mois
    today = timezone.now().date()
    nouveaux_clients = Client.objects.filter(
        date_inscription__month=today.month,
        date_inscription__year=today.year
    ).count()
    
    # Liste des clients par niveau
    clients_vip = Client.objects.filter(niveau_fidelite='VIP').count()
    clients_gold = Client.objects.filter(niveau_fidelite='GOLD').count()
    clients_silver = Client.objects.filter(niveau_fidelite='SILVER').count()
    clients_tous = Client.objects.filter(niveau_fidelite='TOUS').count()
    
    # Liste des clients
    clients = Client.objects.all().order_by('-points_fidelite')
    
    # Activités récentes RÉELLES
    activites = []
    
    # Nouveaux clients récents
    clients_recents = Client.objects.order_by('-date_inscription')[:5]
    for client in clients_recents:
        jours = (timezone.now().date() - client.date_inscription.date()).days
        activites.append({
            'icon': '👤',
            'titre': f'Nouveau client: {client.nom} {client.prenom}',
            'description': f'Niveau: {client.get_niveau_fidelite_display()}',
            'date': client.date_inscription
        })
    
    # Promotions actives récentes
    promotions_recentes = Promotion.objects.filter(est_active=True).order_by('-date_creation')[:3]
    for promo in promotions_recentes:
        activites.append({
            'icon': '%',
            'titre': f'Promotion: {promo.titre}',
            'description': f'Réduction de {promo.reduction}%',
            'date': promo.date_creation
        })
    
    # Réclamations récentes
    reclamations_recentes = Reclamation.objects.order_by('-date_creation')[:2]
    for reclamation in reclamations_recentes:
        activites.append({
            'icon': '⚠️',
            'titre': f'Réclamation de {reclamation.client.nom}',
            'description': reclamation.sujet,
            'date': reclamation.date_creation
        })
    
    # Liste des promotions
    promotions = Promotion.objects.filter(est_active=True)
    for promo in promotions:
        promo.produits_count = promo.produits.count()
    
    # Calcul des pourcentages pour le graphique
    total_clients = max(clients_vip + clients_gold + clients_silver + clients_tous, 1)
    vip_percentage = int((clients_vip / total_clients) * 200)
    gold_percentage = int((clients_gold / total_clients) * 200)
    silver_percentage = int((clients_silver / total_clients) * 200)
    
    context = {
        'clients_fideles': clients_fideles,
        'points_distribues': points_distribues,
        'promotions_actives': promotions_actives,
        'nouveaux_clients': nouveaux_clients,
        'clients_vip': clients_vip,
        'clients_gold': clients_gold,
        'clients_silver': clients_silver,
        'vip_percentage': vip_percentage,
        'gold_percentage': gold_percentage,
        'silver_percentage': silver_percentage,
        'activites': activites,
        'promotions': promotions,
    }
    
    return render(request, 'dashboard/marketing.html', context)

# Tableau de bord Analytics
@login_required
def dashboard_analytics(request):
    # Vérifier les permissions
    if request.user.role != 'ANALYSTE':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour accéder au module Analytics.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # KPI principaux RÉELS
    ventes_mois = Vente.objects.filter(
        date_vente__month=current_month,
        date_vente__year=current_year
    )
    ca_mois = ventes_mois.aggregate(total=Sum('montant_final'))['total'] or 0
    nb_transactions = ventes_mois.count()
    panier_moyen = ca_mois / nb_transactions if nb_transactions > 0 else 0
    
    # Taux de conversion RÉEL (transactions / nombre de clients)
    nb_clients = Client.objects.count() or 1
    taux_conversion = round((nb_transactions / nb_clients) * 100, 1)
    
    # Satisfaction client RÉELLE (basée sur réclamations)
    reclamations = Reclamation.objects.filter(date_creation__month=current_month).count()
    satisfaction = round(5 - (reclamations / max(nb_clients, 1)) * 5, 1)
    
    # Évolution des ventes
    ventes_evolution = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        ca = Vente.objects.filter(date_vente__date=date).aggregate(
            total=Sum('montant_final')
        )['total'] or 0
        ventes_evolution.append({
            'jour': date.strftime('%a'),
            'ca': float(ca)
        })
    
    # Performance globale (gauge)
    performance = 85
    
    # Top produits vendus
    top_produits = []
    lignes = LigneVente.objects.filter(
        vente__date_vente__month=current_month
    ).values(
        'produit__nom'
    ).annotate(
        total_unites=Sum('quantite')
    ).order_by('-total_unites')[:4]
    
    for ligne in lignes:
        top_produits.append({
            'nom': ligne['produit__nom'],
            'unites': ligne['total_unites']
        })
    
    # Répartition par catégorie (pie chart)
    categories = LigneVente.objects.filter(
        vente__date_vente__month=current_month
    ).values(
        'produit__categorie'
    ).annotate(
        total=Sum('montant_ligne')
    )
    
    categories_data = []
    for cat in categories:
        categories_data.append({
            'categorie': cat['produit__categorie'],
            'total': float(cat['total'])
        })
    
    # Moyens de paiement
    moyens_paiement = Vente.objects.filter(
        date_vente__month=current_month
    ).values('moyen_paiement').annotate(
        total=Count('id')
    )
    
    paiement_data = []
    total_ventes = Vente.objects.filter(date_vente__month=current_month).count()
    
    for mp in moyens_paiement:
        pourcentage = (mp['total'] / total_ventes * 100) if total_ventes > 0 else 0
        paiement_data.append({
            'moyen': mp['moyen_paiement'],
            'pourcentage': round(pourcentage, 0)
        })
    
    # Ajouter des informations aux top produits RÉELS (tendance, marge, stock)
    for produit_data in top_produits:
        try:
            produit = Produit.objects.get(nom=produit_data['produit__nom'])
            produit_data['marge'] = round(
                ((float(produit.prix_unitaire) - float(produit.prix_achat)) / float(produit.prix_unitaire)) * 100,
                1
            )
            produit_data['tendance'] = 0  # À calculer avec historique
            produit_data['stock_actuel'] = produit.stock_actuel
        except:
            produit_data['marge'] = 0
            produit_data['tendance'] = 0
            produit_data['stock_actuel'] = 0
    
    context = {
        'ca_mois': float(ca_mois) / 1000000,  # Convertir en millions
        'nb_transactions': nb_transactions,
        'taux_conversion': taux_conversion,
        'satisfaction': satisfaction,
        'ventes_evolution': ventes_evolution,
        'top_produits': top_produits,
        'categories_data': categories_data,
    }
    
    return render(request, 'dashboard/analytics.html', context)


# Vue pour ajouter un produit (accès STOCK uniquement)
@login_required
def stock_add_product(request):
    # Vérifier que l'utilisateur a accès au module Stock
    if request.user.role != 'STOCK':
        messages.error(request, "Accès refusé. Vous n'avez pas les permissions pour gérer les produits.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        nom = request.POST.get('nom')
        reference = request.POST.get('reference')
        categorie = request.POST.get('categorie')
        prix_achat = request.POST.get('prix_achat')
        prix_vente = request.POST.get('prix_vente')
        stock = request.POST.get('stock')
        fournisseur = request.POST.get('fournisseur', '')
        description = request.POST.get('description', '')
        
        # Vérifier que la référence n'existe pas déjà
        if Produit.objects.filter(reference=reference).exists():
            messages.error(request, f"La référence '{reference}' existe déjà. Veuillez en choisir une autre.")
            return render(request, 'dashboard/stock_add_product.html')
        
        try:
            # Créer le produit
            produit = Produit.objects.create(
                nom=nom,
                reference=reference,
                categorie=categorie,
                prix_achat=Decimal(prix_achat),
                prix_unitaire=Decimal(prix_vente),
                stock_actuel=int(stock),
                fournisseur=fournisseur,
                description=description if description else None
            )
            
            # Gestion de l'image si elle est fournie
            if 'image' in request.FILES:
                produit.image = request.FILES['image']
                produit.save()
            
            messages.success(request, f"✅ Produit '{nom}' ajouté avec succès ! Stock initial: {stock} unités")
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du produit: {str(e)}")
            return render(request, 'dashboard/stock_add_product.html')
    
    return render(request, 'dashboard/stock_add_product.html')


# Vue pour modifier un produit (accès STOCK uniquement)
@login_required
def stock_produit_edit(request, produit_id):
    # Vérifier que l'utilisateur est STOCK
    if request.user.role != 'STOCK':
        messages.error(request, "Accès non autorisé. Seul le gestionnaire de stock peut modifier les produits.")
        return redirect('dashboard')
    
    # Récupérer le produit
    try:
        produit = Produit.objects.get(id=produit_id)
    except Produit.DoesNotExist:
        messages.error(request, "Produit introuvable.")
        return redirect('dashboard_stock')
    
    # Récupérer les fournisseurs actifs
    fournisseurs = Fournisseur.objects.filter(est_actif=True).order_by('nom')
    
    if request.method == 'POST':
        try:
            # Récupération des données
            nom = request.POST.get('nom', '').strip()
            reference = request.POST.get('reference', '').strip()
            categorie = request.POST.get('categorie', '').strip()
            prix_achat = request.POST.get('prix_achat', '0').strip()
            prix_unitaire = request.POST.get('prix_unitaire', '0').strip()
            stock_actuel = request.POST.get('stock_actuel', '0').strip()
            stock_critique = request.POST.get('stock_critique', '10').strip()
            seuil_reapprovisionnement = request.POST.get('seuil_reapprovisionnement', '20').strip()
            stock_minimum = request.POST.get('stock_minimum', '5').strip()
            stock_maximum = request.POST.get('stock_maximum', '500').strip()
            description = request.POST.get('description', '').strip()
            code_barre = request.POST.get('code_barre', '').strip()
            fournisseur_id = request.POST.get('fournisseur', '').strip()
            
            # Validations
            if not nom or not reference or not categorie:
                messages.error(request, "Le nom, la référence et la catégorie sont obligatoires.")
                return render(request, 'dashboard/stock_produit_edit.html', {
                    'produit': produit,
                    'fournisseurs': fournisseurs,
                    'CATEGORIES': Produit.CATEGORIES
                })
            
            # Vérifier que la référence n'existe pas déjà (sauf pour le produit actuel)
            if Produit.objects.filter(reference=reference).exclude(id=produit_id).exists():
                messages.error(request, f"Un produit avec la référence '{reference}' existe déjà.")
                return render(request, 'dashboard/stock_produit_edit.html', {
                    'produit': produit,
                    'fournisseurs': fournisseurs,
                    'CATEGORIES': Produit.CATEGORIES
                })
            
            # Récupérer le fournisseur
            fournisseur = None
            if fournisseur_id:
                try:
                    fournisseur = Fournisseur.objects.get(id=int(fournisseur_id))
                except (ValueError, Fournisseur.DoesNotExist):
                    pass
            
            # Gestion de l'image
            if 'image' in request.FILES:
                # Supprimer l'ancienne image si elle existe
                if produit.image:
                    try:
                        old_image_path = produit.image.path
                        if os.path.exists(old_image_path):
                            os.remove(old_image_path)
                    except Exception as e:
                        print(f"Erreur lors de la suppression de l'ancienne image: {e}")
                
                # Sauvegarder la nouvelle image
                produit.image = request.FILES['image']
            
            # Calculer le stock avant modification (pour historique)
            stock_ancien = produit.stock_actuel
            
            # Mettre à jour les champs du produit
            produit.nom = nom
            produit.reference = reference
            produit.categorie = categorie
            produit.prix_achat = float(prix_achat)
            produit.prix_unitaire = float(prix_unitaire)
            produit.stock_actuel = int(stock_actuel)
            produit.stock_critique = int(stock_critique)
            produit.seuil_reapprovisionnement = int(seuil_reapprovisionnement)
            produit.stock_minimum = int(stock_minimum)
            produit.stock_maximum = int(stock_maximum)
            produit.description = description if description else ''
            produit.code_barre = code_barre if code_barre else ''
            produit.fournisseur_principal = fournisseur
            
            produit.save()
            
            # Créer un mouvement de stock si la quantité a changé
            if stock_actuel != stock_ancien:
                difference = int(stock_actuel) - stock_ancien
                type_mouvement = 'ENTREE' if difference > 0 else 'SORTIE'
                
                MouvementStock.objects.create(
                    produit=produit,
                    type_mouvement=type_mouvement,
                    quantite=abs(difference),
                    stock_avant=stock_ancien,
                    stock_apres=int(stock_actuel),
                    raison=f"Ajustement manuel (modification produit)",
                    employe=request.user
                )
            
            messages.success(request, f"✅ Produit '{nom}' modifié avec succès !")
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification du produit: {str(e)}")
    
    context = {
        'produit': produit,
        'fournisseurs': fournisseurs,
        'CATEGORIES': Produit.CATEGORIES
    }
    return render(request, 'dashboard/stock_produit_edit.html', context)


# Vue pour créer un employé (accès RH uniquement)
@login_required
def rh_create_employee(request):
    # Vérifier que l'utilisateur est RH
    if request.user.role != 'RH':
        messages.error(request, "Accès non autorisé. Seul le RH peut créer des employés.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Récupérer les données du formulaire (PAS de username/password)
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        role = request.POST.get('role')
        
        # Validation
        if not all([email, first_name, last_name, role]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return render(request, 'dashboard/rh_create_employee.html')
        
        # Créer l'employé
        try:
            # Générer mot de passe temporaire aléatoire
            import random
            import string
            password_temporaire = 'Carrefour2025!' + ''.join(random.choices(string.digits, k=4))
            
            # Le username sera auto-généré (EMP001, EMP002...) par le modèle
            employe = Employe.objects.create_user(
                username='temp_' + str(timezone.now().timestamp()),  # Temporaire, sera remplacé par save()
                email=email,
                password=password_temporaire,
                first_name=first_name,
                last_name=last_name,
                role=role
            )
            
            # Configurer les horaires de travail (valeurs par défaut ou personnalisées)
            heure_debut = request.POST.get('heure_debut_travail', '08:00')
            heure_fin = request.POST.get('heure_fin_travail', '17:00')
            duree_pause = request.POST.get('duree_pause', '90')
            
            employe.heure_debut_travail = heure_debut
            employe.heure_fin_travail = heure_fin
            employe.duree_pause = int(duree_pause)
            
            # Le rôle détermine automatiquement les accès
            # Le save() va générer employee_id et corriger username
            employe.save()
            
            messages.success(request, f"""
                ✅ Employé créé avec succès !
                
                🆔 ID Employé: {employe.employee_id}
                👤 Identifiant: {employe.username}
                🔑 Mot de passe temporaire: {password_temporaire}
                
                ⚠️ L'employé doit changer son mot de passe lors de sa première connexion.
            """)
            return redirect('dashboard_rh')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création de l'employé: {str(e)}")
            return render(request, 'dashboard/rh_create_employee.html')
    
    return render(request, 'dashboard/rh_create_employee.html')


# Vue liste des employés (RH)
@login_required
def rh_employees_list(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    # Exclure les comptes système protégés (DG, DAF, RH)
    employes = Employe.objects.filter(est_compte_systeme=False).order_by('-date_embauche')
    
    # Filtres
    departement_filtre = request.GET.get('departement', '')
    role_filtre = request.GET.get('role', '')
    statut_filtre = request.GET.get('statut', '')
    
    if departement_filtre:
        employes = employes.filter(departement=departement_filtre)
    if role_filtre:
        employes = employes.filter(role=role_filtre)
    if statut_filtre == 'actif':
        employes = employes.filter(est_actif=True)
    elif statut_filtre == 'inactif':
        employes = employes.filter(est_actif=False)
    
    # Enrichir chaque employé avec le temps depuis la dernière connexion
    from datetime import datetime, timedelta
    now = timezone.now()
    
    for employe in employes:
        if employe.derniere_connexion_custom:
            diff = now - employe.derniere_connexion_custom
            if diff < timedelta(hours=1):
                employe.temps_depuis_connexion = 'recent'  # < 1h
            elif diff < timedelta(days=1):
                employe.temps_depuis_connexion = 'today'  # aujourd'hui
            elif diff < timedelta(days=7):
                employe.temps_depuis_connexion = 'week'  # cette semaine
            else:
                employe.temps_depuis_connexion = 'old'  # plus ancien
        else:
            employe.temps_depuis_connexion = 'never'
    
    # Statistiques
    total_employes = employes.count()
    actifs = employes.filter(est_actif=True).count()
    connectes_aujourd_hui = employes.filter(
        derniere_connexion_custom__gte=timezone.now().replace(hour=0, minute=0, second=0)
    ).count()
    
    # Grouper par département
    from django.db.models import Count
    par_departement = employes.values('departement').annotate(count=Count('id'))
    
    stats = {
        'total_employes': total_employes,
        'actifs': actifs,
        'connectes_aujourd_hui': connectes_aujourd_hui,
        'par_departement': par_departement,
    }
    
    context = {
        'employes': employes,
        'stats': stats,
        'departements': Employe.DEPARTEMENTS,
        'roles': Employe.ROLES,
        'departement_filtre': departement_filtre,
        'role_filtre': role_filtre,
        'statut_filtre': statut_filtre,
    }
    
    # Choisir le template selon la vue demandée
    if request.GET.get('view') == 'list':
        return render(request, 'dashboard/rh_employees_list.html', context)
    else:
        return render(request, 'dashboard/rh_employees_modern.html', context)


# Vue modifier un employé (RH)
@login_required
def rh_employee_edit(request, employee_id):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    employe = get_object_or_404(Employe, id=employee_id)
    
    # Protection des comptes système
    if employe.is_system_account():
        messages.error(request, "❌ Ce compte système est protégé et ne peut pas être modifié.")
        return redirect('rh_employees_list')
    
    if request.method == 'POST':
        employe.first_name = request.POST.get('first_name')
        employe.last_name = request.POST.get('last_name')
        employe.email = request.POST.get('email')
        employe.telephone = request.POST.get('telephone')
        employe.departement = request.POST.get('departement')
        employe.role = request.POST.get('role')
        employe.est_actif = request.POST.get('est_actif') == 'on'
        
        # Gestion des horaires de travail
        heure_debut = request.POST.get('heure_debut_travail')
        heure_fin = request.POST.get('heure_fin_travail')
        duree_pause = request.POST.get('duree_pause')
        
        if heure_debut:
            employe.heure_debut_travail = heure_debut
        if heure_fin:
            employe.heure_fin_travail = heure_fin
        if duree_pause:
            employe.duree_pause = int(duree_pause)
        
        employe.save()
        
        messages.success(request, f"✅ Employé {employe.get_full_name()} modifié avec succès!")
        return redirect('rh_employees_list')
    
    context = {'employe': employe}
    return render(request, 'dashboard/rh_employee_edit.html', context)


# Vue supprimer un employé (RH)
@login_required
def rh_employee_delete(request, employee_id):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    employe = get_object_or_404(Employe, id=employee_id)
    
    # Protection des comptes système
    if employe.is_system_account():
        messages.error(request, "❌ Ce compte système est protégé et ne peut pas être supprimé.")
        return redirect('rh_employees_list')
    
    if request.method == 'POST':
        nom_complet = employe.get_full_name()
        employe.delete()
        messages.success(request, f"✅ Employé {nom_complet} supprimé avec succès!")
        return redirect('rh_employees_list')
    
    context = {'employe': employe}
    return render(request, 'dashboard/rh_employee_delete.html', context)


# Vue gestion des présences (RH)
@login_required
def rh_presences(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    presences = Presence.objects.filter(date=today).select_related('employe')
    employes = Employe.objects.filter(est_actif=True)
    
    # Calculer les statistiques de présence
    total_presents = presences.filter(statut='PRESENT').count()
    total_retards = presences.filter(statut='RETARD').count()
    total_absents = employes.count() - presences.count()
    
    # Calculer les heures travaillées pour chaque présence
    presences_avec_heures = []
    for presence in presences:
        heures_travaillees = presence.calculer_heures_travaillees()
        pourcentage = presence.calculer_pourcentage_presence()
        presences_avec_heures.append({
            'presence': presence,
            'heures_travaillees': round(heures_travaillees, 2),
            'pourcentage': round(pourcentage, 1)
        })
    
    context = {
        'presences_data': presences_avec_heures,
        'employes': employes,
        'today': today,
        'total_presents': total_presents,
        'total_retards': total_retards,
        'total_absents': total_absents,
    }
    return render(request, 'dashboard/rh_presences.html', context)


# Vue gestion des congés (RH)
@login_required
def rh_conges(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    conges_en_attente = Conge.objects.filter(statut='EN_ATTENTE').select_related('employe')
    tous_conges = Conge.objects.all().select_related('employe').order_by('-date_debut')
    
    context = {
        'conges_en_attente': conges_en_attente,
        'tous_conges': tous_conges
    }
    return render(request, 'dashboard/rh_conges.html', context)


# Vue approuver/refuser congé (RH)
@login_required
def rh_conge_action(request, conge_id, action):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    conge = get_object_or_404(Conge, id=conge_id)
    
    if action == 'approve':
        conge.statut = 'APPROUVE'
        conge.approuve_par = request.user
        messages.success(request, f"✅ Congé de {conge.employe.get_full_name()} approuvé!")
    elif action == 'reject':
        conge.statut = 'REFUSE'
        conge.approuve_par = request.user
        messages.warning(request, f"❌ Congé de {conge.employe.get_full_name()} refusé!")
    
    conge.save()
    return redirect('rh_conges')


# Vue demande de congé (Employé)
@login_required
def employee_request_leave(request):
    if request.method == 'POST':
        type_conge = request.POST.get('type_conge')
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        motif = request.POST.get('motif')
        
        # Validation
        if not all([type_conge, date_debut, date_fin, motif]):
            messages.error(request, "❌ Tous les champs sont obligatoires!")
            return redirect('employee_request_leave')
        
        # Vérifier dates
        from datetime import datetime
        debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        if fin < debut:
            messages.error(request, "❌ La date de fin doit être après la date de début!")
            return redirect('employee_request_leave')
        
        if debut < timezone.now().date():
            messages.error(request, "❌ La date de début ne peut pas être dans le passé!")
            return redirect('employee_request_leave')
        
        # Créer la demande
        conge = Conge.objects.create(
            employe=request.user,
            type_conge=type_conge,
            date_debut=debut,
            date_fin=fin,
            motif=motif,
            statut='EN_ATTENTE'
        )
        
        # Calculer nombre de jours
        nb_jours = (fin - debut).days + 1
        
        messages.success(request, f"✅ Votre demande de congé ({nb_jours} jours) a été envoyée! Vous recevrez une notification dès qu'elle sera traitée.")
        return redirect('employee_request_leave')
    
    # GET - Afficher le formulaire et l'historique
    mes_conges = Conge.objects.filter(employe=request.user).order_by('-date_demande')
    
    # Calculer solde de congés (25 jours/an standard en Côte d'Ivoire)
    conges_annee_en_cours = Conge.objects.filter(
        employe=request.user,
        type_conge='ANNUEL',
        statut='APPROUVE',
        date_debut__year=timezone.now().year
    )
    
    jours_pris = sum((c.date_fin - c.date_debut).days + 1 for c in conges_annee_en_cours)
    solde_conges = 25 - jours_pris  # 25 jours standard
    
    context = {
        'mes_conges': mes_conges,
        'solde_conges': solde_conges,
        'jours_pris': jours_pris,
        'types_conge': Conge.TYPES,
    }
    return render(request, 'dashboard/employee_request_leave.html', context)


# Vue calendrier des congés (RH)
@login_required
def rh_conges_calendar(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    # Récupérer tous les congés approuvés pour l'année en cours
    from datetime import datetime
    year = int(request.GET.get('year', timezone.now().year))
    
    conges_approuves = Conge.objects.filter(
        statut='APPROUVE',
        date_debut__year=year
    ).select_related('employe').order_by('date_debut')
    
    # Préparer les données pour le calendrier
    conges_data = []
    for conge in conges_approuves:
        conges_data.append({
            'employe': conge.employe.get_full_name(),
            'type': conge.get_type_conge_display(),
            'date_debut': conge.date_debut.strftime('%Y-%m-%d'),
            'date_fin': conge.date_fin.strftime('%Y-%m-%d'),
            'nb_jours': (conge.date_fin - conge.date_debut).days + 1,
        })
    
    context = {
        'conges_data': conges_data,
        'year': year,
        'years': range(2020, 2031),  # 2020-2030
    }
    return render(request, 'dashboard/rh_conges_calendar.html', context)


# Vue gestion des formations (RH)
@login_required
def rh_formations(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    formations = Formation.objects.all().order_by('-date_debut')
    
    context = {'formations': formations}
    return render(request, 'dashboard/rh_formations.html', context)


# Vue planifications (RH)
@login_required
def rh_planifications(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    employes = Employe.objects.filter(est_actif=True).order_by('departement', 'last_name')
    
    context = {'employes': employes}
    return render(request, 'dashboard/rh_planifications.html', context)


# Vue historique des présences (RH) - NOUVEAU
@login_required
def rh_historique_presences(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    from datetime import datetime, timedelta
    from django.db.models import Count, Q, Avg
    import calendar
    
    # Paramètres
    periode = request.GET.get('periode', 'semaine')
    employe_filtre = request.GET.get('employe', '')
    departement_filtre = request.GET.get('departement', '')
    date_debut_param = request.GET.get('date_debut', '')
    date_fin_param = request.GET.get('date_fin', '')
    
    # Définir les dates selon la période
    now = timezone.now()
    
    if date_debut_param and date_fin_param:
        date_debut = datetime.strptime(date_debut_param, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_param, '%Y-%m-%d').date()
    elif periode == 'jour':
        date_debut = now.date()
        date_fin = now.date()
    elif periode == 'semaine':
        # Du lundi au dimanche de cette semaine
        date_debut = now.date() - timedelta(days=now.weekday())
        date_fin = date_debut + timedelta(days=6)
    elif periode == 'mois':
        date_debut = now.date().replace(day=1)
        # Dernier jour du mois
        _, last_day = calendar.monthrange(now.year, now.month)
        date_fin = now.date().replace(day=last_day)
    elif periode == 'annee':
        date_debut = now.date().replace(month=1, day=1)
        date_fin = now.date().replace(month=12, day=31)
    else:
        date_debut = now.date() - timedelta(days=7)
        date_fin = now.date()
    
    # Récupérer les présences
    presences = Presence.objects.filter(
        date__gte=date_debut,
        date__lte=date_fin
    ).select_related('employe').order_by('-date', 'employe__last_name')
    
    # Appliquer les filtres
    if employe_filtre:
        presences = presences.filter(employe_id=employe_filtre)
    if departement_filtre:
        presences = presences.filter(employe__departement=departement_filtre)
    
    # Calculer les statistiques
    total_presences = presences.count()
    presents = presences.filter(statut='PRESENT').count()
    absents = presences.filter(statut='ABSENT').count()
    retards = presences.filter(statut='RETARD').count()
    
    taux_presence = round((presents / total_presences * 100) if total_presences > 0 else 0, 1)
    
    stats = {
        'presents': presents,
        'absents': absents,
        'retards': retards,
        'taux_presence': taux_presence,
    }
    
    # Pour la vue calendrier (mois)
    calendrier = []
    if periode == 'mois':
        mois_actuel = now.month
        annee_actuel = now.year
        mois_param = request.GET.get('mois', f'{annee_actuel}-{mois_actuel:02d}')
        
        try:
            annee, mois = map(int, mois_param.split('-'))
        except:
            annee, mois = annee_actuel, mois_actuel
        
        # Premier jour du mois
        premier_jour = datetime(annee, mois, 1).date()
        _, nb_jours = calendar.monthrange(annee, mois)
        
        # Décalage pour le premier jour (lundi = 0)
        premier_jour_semaine = premier_jour.weekday()
        
        # Ajouter des jours vides au début
        for _ in range(premier_jour_semaine):
            calendrier.append({'numero': '', 'statut': '', 'presents': 0, 'absents': 0})
        
        # Ajouter les jours du mois
        for jour in range(1, nb_jours + 1):
            date_jour = datetime(annee, mois, jour).date()
            presences_jour = Presence.objects.filter(date=date_jour)
            
            if departement_filtre:
                presences_jour = presences_jour.filter(employe__departement=departement_filtre)
            
            presents_jour = presences_jour.filter(Q(statut='PRESENT') | Q(statut='RETARD')).count()
            absents_jour = presences_jour.filter(statut='ABSENT').count()
            
            statut = ''
            if presents_jour > absents_jour and presents_jour > 0:
                statut = 'present'
            elif absents_jour > 0:
                statut = 'absent'
            
            calendrier.append({
                'numero': jour,
                'statut': statut,
                'presents': presents_jour,
                'absents': absents_jour,
            })
        
        # Infos pour navigation
        mois_nom = calendar.month_name[mois]
        mois_precedent = f'{annee}-{mois-1:02d}' if mois > 1 else f'{annee-1}-12'
        mois_suivant = f'{annee}-{mois+1:02d}' if mois < 12 else f'{annee+1}-01'
    else:
        mois_nom = ''
        annee = now.year
        mois_precedent = ''
        mois_suivant = ''
        mois_actuel = f'{now.year}-{now.month:02d}'
    
    # Liste des employés pour le filtre
    employes_list = Employe.objects.filter(est_actif=True, est_compte_systeme=False).order_by('last_name')
    
    context = {
        'presences': presences,
        'stats': stats,
        'periode': periode,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'employe_filtre': employe_filtre,
        'departement_filtre': departement_filtre,
        'employes_list': employes_list,
        'departements': Employe.DEPARTEMENTS,
        'calendrier': calendrier,
        'mois_nom': mois_nom,
        'annee': annee,
        'mois_precedent': mois_precedent,
        'mois_suivant': mois_suivant,
        'mois_actuel': mois_actuel,
    }
    
    # Export Excel
    if request.GET.get('export') == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Présences {periode}"
        
        # En-têtes
        headers = ['Employé', 'ID', 'Département', 'Date', 'Heure Arrivée', 'Heure Départ', 'Temps Travaillé', 'Statut']
        ws.append(headers)
        
        # Style des en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for presence in presences:
            ws.append([
                presence.employe.get_full_name(),
                presence.employe.employee_id,
                presence.employe.get_departement_display(),
                presence.date.strftime('%d/%m/%Y'),
                presence.heure_arrivee.strftime('%H:%M') if presence.heure_arrivee else '',
                presence.heure_depart.strftime('%H:%M') if presence.heure_depart else '',
                str(presence.temps_travaille_total) if presence.temps_travaille_total else '',
                presence.get_statut_display(),
            ])
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Créer la réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=presences_{periode}_{date_debut}_{date_fin}.xlsx'
        wb.save(response)
        return response
    
    return render(request, 'dashboard/rh_historique_presences.html', context)


# Vue historique RH complet
@login_required
def rh_historique(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    from datetime import datetime, timedelta
    from django.db.models import Q
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.core.paginator import Paginator
    
    # Filtres
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    employe_id = request.GET.get('employe', '')
    type_action = request.GET.get('type', '')
    
    # Dates par défaut (30 derniers jours)
    if not date_debut:
        date_debut = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = timezone.now().strftime('%Y-%m-%d')
    
    # Conversion en objets date
    date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
    date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
    
    # Aggrégation des actions
    actions = []
    
    # 1. Créations d'employés
    if not type_action or type_action == 'CREATION_EMPLOYE':
        employes_crees = Employe.objects.filter(
            date_embauche__range=[date_debut_obj, date_fin_obj]
        )
        if employe_id:
            employes_crees = employes_crees.filter(id=employe_id)
        
        for emp in employes_crees:
            actions.append({
                'type': 'CREATION_EMPLOYE',
                'type_label': '➕ Création Employé',
                'date': timezone.make_aware(datetime.combine(emp.date_embauche, datetime.min.time())),
                'employe': emp,
                'details': f'{emp.get_full_name()} - {emp.get_role_display()} ({emp.employee_id})',
                'auteur': 'Système RH'
            })
    
    # 2. Congés (demandes et approbations)
    if not type_action or type_action == 'CONGE':
        conges = Conge.objects.filter(
            date_demande__range=[
                timezone.make_aware(datetime.combine(date_debut_obj, datetime.min.time())),
                timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
            ]
        ).select_related('employe', 'approuve_par')
        if employe_id:
            conges = conges.filter(employe_id=employe_id)
        
        for conge in conges:
            nb_jours = (conge.date_fin - conge.date_debut).days + 1
            actions.append({
                'type': 'CONGE',
                'type_label': f'🏖️ Congé ({conge.get_statut_display()})',
                'date': conge.date_demande,
                'employe': conge.employe,
                'details': f'{conge.get_type_conge_display()} - {conge.date_debut.strftime("%d/%m")} au {conge.date_fin.strftime("%d/%m")} ({nb_jours}j)',
                'auteur': conge.approuve_par.get_full_name() if conge.approuve_par else '-'
            })
    
    # 3. Formations
    if not type_action or type_action == 'FORMATION':
        formations = Formation.objects.filter(
            date_debut__range=[date_debut_obj, date_fin_obj]
        ).prefetch_related('participants')
        
        for formation in formations:
            participants = formation.participants.all()
            if employe_id:
                participants = participants.filter(id=employe_id)
            
            for participant in participants:
                actions.append({
                    'type': 'FORMATION',
                    'type_label': '🎓 Formation',
                    'date': timezone.make_aware(datetime.combine(formation.date_debut, datetime.min.time())),
                    'employe': participant,
                    'details': f'{formation.titre} ({formation.date_debut.strftime("%d/%m")} - {formation.date_fin.strftime("%d/%m")})',
                    'auteur': 'Direction RH'
                })
    
    # 4. Présences
    if not type_action or type_action == 'PRESENCE':
        presences = Presence.objects.filter(
            date__range=[date_debut_obj, date_fin_obj]
        ).select_related('employe')
        if employe_id:
            presences = presences.filter(employe_id=employe_id)
        
        for presence in presences:
            statut_icon = '✅' if presence.statut == 'PRESENT' else '❌' if presence.statut == 'ABSENT' else '🕐'
            actions.append({
                'type': 'PRESENCE',
                'type_label': f'{statut_icon} Présence',
                'date': timezone.make_aware(datetime.combine(presence.date, datetime.min.time())),
                'employe': presence.employe,
                'details': f'{presence.get_statut_display()} - {presence.heure_arrivee or "N/A"} à {presence.heure_depart or "N/A"}',
                'auteur': 'Système Pointage'
            })
    
    # 5. Réclamations
    if not type_action or type_action == 'RECLAMATION':
        reclamations = Reclamation.objects.filter(
            date_soumission__range=[
                timezone.make_aware(datetime.combine(date_debut_obj, datetime.min.time())),
                timezone.make_aware(datetime.combine(date_fin_obj, datetime.max.time()))
            ]
        ).select_related('employe')
        if employe_id:
            reclamations = reclamations.filter(employe_id=employe_id)
        
        for reclamation in reclamations:
            actions.append({
                'type': 'RECLAMATION',
                'type_label': f'📢 Réclamation ({reclamation.get_statut_display()})',
                'date': reclamation.date_soumission,
                'employe': reclamation.employe,
                'details': reclamation.sujet,
                'auteur': '-'
            })
    
    # Tri chronologique inverse
    actions.sort(key=lambda x: x['date'], reverse=True)
    
    # Export Excel si demandé
    if request.GET.get('export') == 'excel':
        # Créer workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historique RH"
        
        # En-têtes
        headers = ['DATE', 'HEURE', 'TYPE ACTION', 'EMPLOYÉ', 'DÉTAILS', 'AUTEUR']
        ws.append(headers)
        
        # Style en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Données
        for action in actions:
            ws.append([
                action['date'].strftime('%d/%m/%Y'),
                action['date'].strftime('%H:%M'),
                action['type_label'],
                action['employe'].get_full_name(),
                action['details'],
                action['auteur']
            ])
        
        # Ajuster largeurs colonnes
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Historique_RH_{date_debut}_{date_fin}.xlsx'
        wb.save(response)
        return response
    
    # Pagination (50 items par page)
    paginator = Paginator(actions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Liste des employés pour filtre
    tous_employes = Employe.objects.filter(est_actif=True).order_by('first_name')
    
    context = {
        'page_obj': page_obj,
        'tous_employes': tous_employes,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'employe_id': employe_id,
        'type_action': type_action,
        'total_actions': len(actions),
        'types_actions': [
            ('', 'Tous les types'),
            ('CREATION_EMPLOYE', '➕ Création Employé'),
            ('CONGE', '🏖️ Congé'),
            ('FORMATION', '🎓 Formation'),
            ('PRESENCE', '✅ Présence'),
            ('RECLAMATION', '📢 Réclamation'),
        ]
    }
    return render(request, 'dashboard/rh_historique.html', context)


# Vue ajouter présence (RH)
@login_required
def rh_presence_add(request):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        employe_id = request.POST.get('employe_id')
        date = request.POST.get('date')
        heure_premiere_arrivee = request.POST.get('heure_premiere_arrivee')
        heure_derniere_depart = request.POST.get('heure_derniere_depart')
        motif_absence = request.POST.get('motif_absence', '')
        tolerance_retard = request.POST.get('tolerance_retard', 60)
        
        try:
            employe = Employe.objects.get(id=employe_id)
            presence = Presence.objects.create(
                employe=employe,
                date=date,
                heure_premiere_arrivee=heure_premiere_arrivee if heure_premiere_arrivee else None,
                heure_derniere_depart=heure_derniere_depart if heure_derniere_depart else None,
                motif_absence=motif_absence,
                tolerance_retard=int(tolerance_retard)
            )
            messages.success(request, f"✅ Présence ajoutée pour {employe.get_full_name()}")
        except Exception as e:
            messages.error(request, f"❌ Erreur: {str(e)}")
        
        return redirect('rh_presences')
    
    employes = Employe.objects.filter(est_actif=True).order_by('last_name')
    context = {'employes': employes, 'today': timezone.now().date()}
    return render(request, 'dashboard/rh_presence_form.html', context)


# Vue modifier présence (RH)
@login_required
def rh_presence_edit(request, presence_id):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    presence = get_object_or_404(Presence, id=presence_id)
    
    if request.method == 'POST':
        presence.date = request.POST.get('date')
        presence.heure_premiere_arrivee = request.POST.get('heure_premiere_arrivee') or None
        presence.heure_derniere_depart = request.POST.get('heure_derniere_depart') or None
        presence.motif_absence = request.POST.get('motif_absence', '')
        presence.tolerance_retard = int(request.POST.get('tolerance_retard', 60))
        
        try:
            presence.save()
            messages.success(request, f"✅ Présence mise à jour pour {presence.employe.get_full_name()}")
            return redirect('rh_presences')
        except Exception as e:
            messages.error(request, f"❌ Erreur: {str(e)}")
    
    context = {'presence': presence}
    return render(request, 'dashboard/rh_presence_form.html', context)


# Vue supprimer présence (RH)
@login_required
def rh_presence_delete(request, presence_id):
    if request.user.role != 'RH':
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    presence = get_object_or_404(Presence, id=presence_id)
    
    if request.method == 'POST':
        employe_name = presence.employe.get_full_name()
        presence.delete()
        messages.success(request, f"✅ Présence de {employe_name} supprimée")
        return redirect('rh_presences')
    
    context = {'presence': presence}
    return render(request, 'dashboard/rh_presence_delete.html', context)


# ============================================
# GESTION DES FOURNISSEURS (SPRINT 1 - JOUR 5-6)
# ============================================

@login_required
def stock_fournisseurs_list(request):
    """Liste tous les fournisseurs avec recherche et filtrage"""
    if request.user.role not in ['STOCK', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "❌ Accès refusé")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    fournisseurs = Fournisseur.objects.all().order_by('-date_creation')
    
    # Recherche
    search = request.GET.get('search', '')
    if search:
        fournisseurs = fournisseurs.filter(
            Q(nom__icontains=search) |
            Q(contact__icontains=search) |
            Q(email__icontains=search) |
            Q(telephone__icontains=search)
        )
    
    # Filtre par statut
    statut = request.GET.get('statut', '')
    if statut == 'actif':
        fournisseurs = fournisseurs.filter(est_actif=True)
    elif statut == 'inactif':
        fournisseurs = fournisseurs.filter(est_actif=False)
    
    # Statistiques
    total_fournisseurs = Fournisseur.objects.count()
    fournisseurs_actifs = Fournisseur.objects.filter(est_actif=True).count()
    fournisseurs_inactifs = Fournisseur.objects.filter(est_actif=False).count()
    
    context = {
        'fournisseurs': fournisseurs,
        'search': search,
        'statut': statut,
        'total_fournisseurs': total_fournisseurs,
        'fournisseurs_actifs': fournisseurs_actifs,
        'fournisseurs_inactifs': fournisseurs_inactifs,
    }
    return render(request, 'dashboard/stock_fournisseurs_list.html', context)


@login_required
def stock_fournisseur_detail(request, fournisseur_id):
    """Affiche les détails d'un fournisseur"""
    if not request.user.has_perm('CarrefourApp.view_fournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de voir ce fournisseur")
        return redirect('dashboard')
    
    fournisseur = get_object_or_404(Fournisseur, id=fournisseur_id)
    
    # Produits fournis
    produits = Produit.objects.filter(fournisseur_principal=fournisseur)
    
    # Commandes
    commandes = CommandeFournisseur.objects.filter(fournisseur=fournisseur).order_by('-date_commande')[:10]
    
    # Statistiques
    total_commandes = fournisseur.nombre_commandes()
    total_produits = fournisseur.nombre_produits()
    montant_total = fournisseur.montant_total_commandes()
    
    context = {
        'fournisseur': fournisseur,
        'produits': produits,
        'commandes': commandes,
        'total_commandes': total_commandes,
        'total_produits': total_produits,
        'montant_total': montant_total,
    }
    return render(request, 'dashboard/stock_fournisseur_detail.html', context)


@login_required
def stock_fournisseur_create(request):
    """Crée un nouveau fournisseur"""
    if not request.user.has_perm('CarrefourApp.add_fournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de créer un fournisseur")
        return redirect('dashboard')
    
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        contact = request.POST.get('contact', '').strip()
        email = request.POST.get('email', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        adresse = request.POST.get('adresse', '').strip()
        delai_livraison = request.POST.get('delai_livraison_moyen', '7')
        conditions_paiement = request.POST.get('conditions_paiement', '').strip()
        est_actif = request.POST.get('est_actif') == 'on'
        
        # Validation
        if not nom:
            messages.error(request, "❌ Le nom du fournisseur est obligatoire")
        elif not contact:
            messages.error(request, "❌ Le contact du fournisseur est obligatoire")
        else:
            try:
                fournisseur = Fournisseur.objects.create(
                    nom=nom,
                    contact=contact,
                    email=email,
                    telephone=telephone,
                    adresse=adresse,
                    delai_livraison_moyen=int(delai_livraison),
                    conditions_paiement=conditions_paiement,
                    est_actif=est_actif
                )
                messages.success(request, f"✅ Fournisseur '{nom}' créé avec succès")
                return redirect('stock_fournisseur_detail', fournisseur_id=fournisseur.id)
            except Exception as e:
                messages.error(request, f"❌ Erreur lors de la création: {str(e)}")
    
    context = {
        'action': 'create',
        'page_title': 'Nouveau Fournisseur'
    }
    return render(request, 'dashboard/stock_fournisseur_form.html', context)


@login_required
def stock_fournisseur_edit(request, fournisseur_id):
    """Modifie un fournisseur existant"""
    if not request.user.has_perm('CarrefourApp.change_fournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de modifier ce fournisseur")
        return redirect('dashboard')
    
    fournisseur = get_object_or_404(Fournisseur, id=fournisseur_id)
    
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        contact = request.POST.get('contact', '').strip()
        email = request.POST.get('email', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        adresse = request.POST.get('adresse', '').strip()
        delai_livraison = request.POST.get('delai_livraison_moyen', '7')
        conditions_paiement = request.POST.get('conditions_paiement', '').strip()
        est_actif = request.POST.get('est_actif') == 'on'
        
        # Validation
        if not nom:
            messages.error(request, "❌ Le nom du fournisseur est obligatoire")
        elif not contact:
            messages.error(request, "❌ Le contact du fournisseur est obligatoire")
        else:
            try:
                fournisseur.nom = nom
                fournisseur.contact = contact
                fournisseur.email = email
                fournisseur.telephone = telephone
                fournisseur.adresse = adresse
                fournisseur.delai_livraison_moyen = int(delai_livraison)
                fournisseur.conditions_paiement = conditions_paiement
                fournisseur.est_actif = est_actif
                fournisseur.save()
                
                messages.success(request, f"✅ Fournisseur '{nom}' modifié avec succès")
                return redirect('stock_fournisseur_detail', fournisseur_id=fournisseur.id)
            except Exception as e:
                messages.error(request, f"❌ Erreur lors de la modification: {str(e)}")
    
    context = {
        'fournisseur': fournisseur,
        'action': 'edit',
        'page_title': f'Modifier {fournisseur.nom}'
    }
    return render(request, 'dashboard/stock_fournisseur_form.html', context)


@login_required
def stock_fournisseur_delete(request, fournisseur_id):
    """Supprime un fournisseur"""
    if not request.user.has_perm('CarrefourApp.delete_fournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de supprimer ce fournisseur")
        return redirect('dashboard')
    
    fournisseur = get_object_or_404(Fournisseur, id=fournisseur_id)
    
    if request.method == 'POST':
        nom = fournisseur.nom
        
        # Vérifier si le fournisseur a des produits associés
        nb_produits = fournisseur.nombre_produits()
        if nb_produits > 0:
            messages.warning(request, f"⚠️ Ce fournisseur a {nb_produits} produit(s) associé(s). Suppression impossible.")
            return redirect('stock_fournisseur_detail', fournisseur_id=fournisseur.id)
        
        # Vérifier si le fournisseur a des commandes
        nb_commandes = fournisseur.nombre_commandes()
        if nb_commandes > 0:
            messages.warning(request, f"⚠️ Ce fournisseur a {nb_commandes} commande(s) associée(s). Suppression impossible.")
            return redirect('stock_fournisseur_detail', fournisseur_id=fournisseur.id)
        
        fournisseur.delete()
        messages.success(request, f"✅ Fournisseur '{nom}' supprimé avec succès")
        return redirect('stock_fournisseurs_list')
    
    context = {'fournisseur': fournisseur}
    return render(request, 'dashboard/stock_fournisseur_delete.html', context)


# ============================================
# GESTION DES COMMANDES FOURNISSEURS (SPRINT 1 - JOUR 9-10)
# ============================================

@login_required
def stock_commandes_list(request):
    """Liste toutes les commandes fournisseurs"""
    if not request.user.has_perm('CarrefourApp.view_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de voir les commandes")
        return redirect('dashboard')
    
    commandes = CommandeFournisseur.objects.all().select_related('fournisseur', 'employe').order_by('-date_commande')
    
    # Filtres
    statut = request.GET.get('statut', '')
    if statut:
        commandes = commandes.filter(statut=statut)
    
    fournisseur_id = request.GET.get('fournisseur', '')
    if fournisseur_id:
        commandes = commandes.filter(fournisseur_id=fournisseur_id)
    
    search = request.GET.get('search', '')
    if search:
        commandes = commandes.filter(
            Q(numero_commande__icontains=search) |
            Q(fournisseur__nom__icontains=search)
        )
    
    # Statistiques
    total_commandes = CommandeFournisseur.objects.count()
    commandes_en_attente = CommandeFournisseur.objects.filter(statut='EN_ATTENTE').count()
    commandes_validees = CommandeFournisseur.objects.filter(statut='VALIDEE').count()
    commandes_livrees = CommandeFournisseur.objects.filter(statut='LIVREE').count()
    montant_total = CommandeFournisseur.objects.aggregate(total=Sum('montant_total'))['total'] or 0
    
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    
    context = {
        'commandes': commandes,
        'statut': statut,
        'fournisseur_id': fournisseur_id,
        'search': search,
        'total_commandes': total_commandes,
        'commandes_en_attente': commandes_en_attente,
        'commandes_validees': commandes_validees,
        'commandes_livrees': commandes_livrees,
        'montant_total': montant_total,
        'fournisseurs': fournisseurs,
    }
    return render(request, 'dashboard/stock_commandes_list.html', context)


@login_required
def stock_commande_detail(request, commande_id):
    """Affiche les détails d'une commande"""
    if not request.user.has_perm('CarrefourApp.view_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de voir cette commande")
        return redirect('dashboard')
    
    commande = get_object_or_404(CommandeFournisseur, id=commande_id)
    lignes = commande.lignes.all().select_related('produit')  # ✅ CORRECTION
    
    context = {
        'commande': commande,
        'lignes': lignes,
    }
    return render(request, 'dashboard/stock_commande_detail.html', context)


@login_required
def stock_commande_create(request):
    """Crée une nouvelle commande fournisseur"""
    if not request.user.has_perm('CarrefourApp.add_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de créer une commande")
        return redirect('dashboard')
    
    if request.method == 'POST':
        fournisseur_id = request.POST.get('fournisseur')
        date_livraison_prevue = request.POST.get('date_livraison_prevue')
        notes = request.POST.get('notes', '').strip()
        
        # Récupérer les produits et quantités
        produits_ids = request.POST.getlist('produit[]')
        quantites = request.POST.getlist('quantite[]')
        prix_unitaires = request.POST.getlist('prix_unitaire[]')
        
        # Validation
        if not fournisseur_id:
            messages.error(request, "❌ Veuillez sélectionner un fournisseur")
        elif not produits_ids or not any(quantites):
            messages.error(request, "❌ Veuillez ajouter au moins un produit à la commande")
        else:
            try:
                from datetime import datetime
                
                fournisseur = Fournisseur.objects.get(id=fournisseur_id)
                employe = Employe.objects.filter(user=request.user).first()
                
                # Créer la commande
                commande = CommandeFournisseur.objects.create(
                    fournisseur=fournisseur,
                    date_livraison_prevue=datetime.strptime(date_livraison_prevue, '%Y-%m-%d').date() if date_livraison_prevue else None,
                    employe=employe,
                    notes=notes,
                    statut='EN_ATTENTE'
                )
                
                # Créer les lignes de commande
                montant_total = Decimal('0')
                for i, produit_id in enumerate(produits_ids):
                    if produit_id and quantites[i] and int(quantites[i]) > 0:
                        produit = Produit.objects.get(id=produit_id)
                        quantite = int(quantites[i])
                        prix_unitaire = Decimal(prix_unitaires[i]) if prix_unitaires[i] else produit.prix_achat
                        
                        LigneCommandeFournisseur.objects.create(
                            commande=commande,
                            produit=produit,
                            quantite_commandee=quantite,
                            prix_unitaire=prix_unitaire
                        )
                        
                        montant_total += quantite * prix_unitaire
                
                # Mettre à jour le montant total
                commande.montant_total = montant_total
                commande.save()
                
                messages.success(request, f"✅ Commande {commande.numero_commande} créée avec succès")
                return redirect('stock_commande_detail', commande_id=commande.id)
            except Exception as e:
                messages.error(request, f"❌ Erreur lors de la création: {str(e)}")
    
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    produits = Produit.objects.filter(est_actif=True)
    
    context = {
        'fournisseurs': fournisseurs,
        'produits': produits,
    }
    return render(request, 'dashboard/stock_commande_form.html', context)


@login_required
def stock_commande_valider(request, commande_id):
    """Valide une commande"""
    if not request.user.has_perm('CarrefourApp.change_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de valider cette commande")
        return redirect('dashboard')
    
    commande = get_object_or_404(CommandeFournisseur, id=commande_id)
    
    if commande.statut != 'EN_ATTENTE':
        messages.warning(request, "⚠️ Cette commande ne peut pas être validée")
        return redirect('stock_commande_detail', commande_id=commande.id)
    
    commande.statut = 'VALIDEE'
    commande.save()
    
    messages.success(request, f"✅ Commande {commande.numero_commande} validée avec succès")
    return redirect('stock_commande_detail', commande_id=commande.id)


@login_required
def stock_commande_recevoir(request, commande_id):
    """Marque une commande comme reçue/livrée"""
    if not request.user.has_perm('CarrefourApp.change_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission de recevoir cette commande")
        return redirect('dashboard')
    
    commande = get_object_or_404(CommandeFournisseur, id=commande_id)
    
    if request.method == 'POST':
        from datetime import date
        
        # Récupérer les quantités reçues
        lignes = commande.lignes.all()  # ✅ CORRECTION
        
        for ligne in lignes:
            quantite_recue = request.POST.get(f'quantite_recue_{ligne.id}', 0)
            ligne.quantite_recue = int(quantite_recue)
            ligne.save()
            
            # ✅ AUGMENTER LE STOCK DU PRODUIT
            if ligne.quantite_recue > 0:
                produit = ligne.produit
                stock_avant = produit.stock_actuel
                nouveau_stock = stock_avant + ligne.quantite_recue
                Produit.objects.filter(id=produit.id).update(stock_actuel=nouveau_stock)
                
                # Créer mouvement de stock
                MouvementStock.objects.create(
                    produit=produit,
                    type_mouvement='ENTREE',
                    quantite=ligne.quantite_recue,
                    stock_avant=stock_avant,
                    raison=f'Réception commande {commande.numero_commande}',
                    employe=request.user
                )
                print(f"   📦 {produit.nom}: {stock_avant} → {nouveau_stock} (+{ligne.quantite_recue})")
        
        # Marquer comme livrée
        commande.statut = 'LIVREE'
        commande.date_livraison_reelle = date.today()
        commande.save()
        
        messages.success(request, f"✅ Commande {commande.numero_commande} réceptionnée avec succès")
        return redirect('stock_commande_detail', commande_id=commande.id)
    
    lignes = commande.lignes.all().select_related('produit')  # ✅ CORRECTION
    
    context = {
        'commande': commande,
        'lignes': lignes,
    }
    return render(request, 'dashboard/stock_commande_recevoir.html', context)


@login_required
def stock_commande_annuler(request, commande_id):
    """Annule une commande"""
    if not request.user.has_perm('CarrefourApp.change_commandefournisseur'):
        messages.error(request, "❌ Vous n'avez pas la permission d'annuler cette commande")
        return redirect('dashboard')
    
    commande = get_object_or_404(CommandeFournisseur, id=commande_id)
    
    if commande.statut == 'LIVREE':
        messages.warning(request, "⚠️ Une commande livrée ne peut pas être annulée")
        return redirect('stock_commande_detail', commande_id=commande.id)
    
    if request.method == 'POST':
        commande.statut = 'ANNULEE'
        commande.save()
        
        messages.success(request, f"✅ Commande {commande.numero_commande} annulée")
        return redirect('stock_commandes_list')
    
    context = {'commande': commande}
    return render(request, 'dashboard/stock_commande_annuler.html', context)


# ============================================
# GESTION DES ALERTES STOCK (SPRINT 1 - JOUR 11-12)
# ============================================

@login_required
def stock_alertes_list(request):
    """Liste toutes les alertes de stock"""
    if request.user.role not in ['STOCK', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "❌ Accès refusé")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    alertes = AlerteStock.objects.all().select_related('produit').order_by('-date_alerte')
    
    # Filtres
    type_alerte = request.GET.get('type', '')
    if type_alerte:
        alertes = alertes.filter(type_alerte=type_alerte)
    
    statut = request.GET.get('statut', '')
    if statut == 'actives':
        alertes = alertes.filter(est_resolue=False)
    elif statut == 'resolues':
        alertes = alertes.filter(est_resolue=True)
    
    # Statistiques
    total_alertes = AlerteStock.objects.count()
    alertes_actives = AlerteStock.objects.filter(est_resolue=False).count()
    alertes_rupture = AlerteStock.objects.filter(type_alerte='RUPTURE', est_resolue=False).count()
    alertes_critiques = AlerteStock.objects.filter(type_alerte='SEUIL_CRITIQUE', est_resolue=False).count()
    alertes_surstock = AlerteStock.objects.filter(type_alerte='SURSTOCK', est_resolue=False).count()
    
    context = {
        'alertes': alertes,
        'type_alerte': type_alerte,
        'statut': statut,
        'total_alertes': total_alertes,
        'alertes_actives': alertes_actives,
        'alertes_rupture': alertes_rupture,
        'alertes_critiques': alertes_critiques,
        'alertes_surstock': alertes_surstock,
    }
    return render(request, 'dashboard/stock_alertes_list.html', context)


@login_required
def stock_alerte_resoudre(request, alerte_id):
    """Marque une alerte comme résolue"""
    if not request.user.has_perm('CarrefourApp.change_alertestock'):
        messages.error(request, "❌ Vous n'avez pas la permission de résoudre cette alerte")
        return redirect('dashboard')
    
    alerte = get_object_or_404(AlerteStock, id=alerte_id)
    
    from datetime import datetime
    alerte.est_resolue = True
    alerte.date_resolution = datetime.now()
    alerte.save()
    
    messages.success(request, f"✅ Alerte résolue pour le produit {alerte.produit.nom}")
    return redirect('stock_alertes_list')


@login_required
def stock_mouvements_list(request):
    """Liste tous les mouvements de stock"""
    if not request.user.has_perm('CarrefourApp.view_mouvementstock'):
        messages.error(request, "❌ Vous n'avez pas la permission de voir les mouvements")
        return redirect('dashboard')
    
    mouvements = MouvementStock.objects.all().select_related('produit', 'employe').order_by('-date_mouvement')[:100]
    
    # Filtres
    type_mouvement = request.GET.get('type', '')
    if type_mouvement:
        mouvements = MouvementStock.objects.filter(type_mouvement=type_mouvement).select_related('produit', 'employe').order_by('-date_mouvement')[:100]
    
    produit_id = request.GET.get('produit', '')
    if produit_id:
        mouvements = MouvementStock.objects.filter(produit_id=produit_id).select_related('produit', 'employe').order_by('-date_mouvement')[:100]
    
    # Statistiques
    total_mouvements = MouvementStock.objects.count()
    entrees = MouvementStock.objects.filter(type_mouvement='ENTREE').count()
    sorties = MouvementStock.objects.filter(type_mouvement='SORTIE').count()
    ajustements = MouvementStock.objects.filter(type_mouvement='AJUSTEMENT').count()
    
    produits = Produit.objects.filter(est_actif=True)
    
    context = {
        'mouvements': mouvements,
        'type_mouvement': type_mouvement,
        'produit_id': produit_id,
        'total_mouvements': total_mouvements,
        'entrees': entrees,
        'sorties': sorties,
        'ajustements': ajustements,
        'produits': produits,
    }
    return render(request, 'dashboard/stock_mouvements_list.html', context)


# =====================================================
# SPRINT 2 - VUES MODULE CAISSE (POS)
# =====================================================


@login_required
def pos_interface(request):
    """
    Interface principale de la caisse (Point of Sale)
    """
    # Vérifier si une session de caisse est ouverte pour le caissier
    session_ouverte = SessionCaisse.objects.filter(
        caissier=request.user,
        est_cloturee=False
    ).first()
    
    if not session_ouverte:
        # Rediriger vers l'ouverture de session
        messages.warning(request, "Vous devez ouvrir une session de caisse avant de commencer.")
        return redirect('pos_ouvrir_session')
    
    # Récupérer la transaction en cours (si existe)
    transaction_en_cours = Transaction.objects.filter(
        caissier=request.user,
        statut='EN_COURS'
    ).first()
    
    # Produits disponibles (stock > 0)
    produits = Produit.objects.filter(
        est_actif=True,
        stock_actuel__gt=0
    ).order_by('nom')
    
    # Types de paiement actifs
    types_paiement = TypePaiement.objects.filter(est_actif=True)
    
    # Catégories pour filtrage
    categories = Produit.CATEGORIES
    
    context = {
        'session': session_ouverte,
        'transaction': transaction_en_cours,
        'produits': produits,
        'types_paiement': types_paiement,
        'categories': categories,
    }
    
    return render(request, 'caisse/pos_interface.html', context)


@login_required
def pos_nouvelle_transaction(request):
    """
    Crée une nouvelle transaction de vente
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Méthode non autorisée. Utilisez POST.'
        }, status=405)
    
    # Vérifier session ouverte
    session_ouverte = SessionCaisse.objects.filter(
        caissier=request.user,
        est_cloturee=False
    ).first()
    
    if not session_ouverte:
        return JsonResponse({
            'success': False,
            'error': 'Aucune session de caisse ouverte. Veuillez ouvrir une session d\'abord.'
        })
    
    # Annuler toute transaction en cours
    Transaction.objects.filter(
        caissier=request.user,
        statut='EN_COURS'
    ).update(statut='ANNULEE')
    
    # Créer nouvelle transaction
    try:
        transaction = Transaction.objects.create(
            caissier=request.user,
            statut='EN_COURS'
        )
        
        return JsonResponse({
            'success': True,
            'transaction_id': transaction.id,
            'numero_ticket': transaction.numero_ticket
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la création: {str(e)}'
        }, status=500)


@login_required
def pos_ajouter_produit(request):
    """
    Ajoute un produit à la transaction en cours
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Méthode non autorisée. Utilisez POST.'
        }, status=405)
    
    try:
        data = json.loads(request.body)
        produit_id = data.get('produit_id')
        quantite = int(data.get('quantite', 1))
        
        # Transaction en cours
        transaction = Transaction.objects.filter(
            caissier=request.user,
            statut='EN_COURS'
        ).first()
        
        if not transaction:
            return JsonResponse({
                'success': False,
                'error': 'Aucune transaction en cours.'
            })
        
        try:
            produit = Produit.objects.get(id=produit_id)
            
            # Vérifier stock disponible
            if produit.stock_actuel < quantite:
                return JsonResponse({
                    'success': False,
                    'error': f'Stock insuffisant. Disponible: {produit.stock_actuel}'
                })
            
            # Vérifier si le produit existe déjà dans la transaction
            ligne_existante = LigneTransaction.objects.filter(
                transaction=transaction,
                produit=produit
            ).first()
            
            if ligne_existante:
                # Augmenter la quantité
                nouvelle_quantite = ligne_existante.quantite + quantite
                if produit.stock_actuel < nouvelle_quantite:
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuffisant. Disponible: {produit.stock_actuel}'
                    })
                ligne_existante.quantite = nouvelle_quantite
                ligne_existante.save()
                ligne = ligne_existante
            else:
                # Créer nouvelle ligne
                ligne = LigneTransaction.objects.create(
                    transaction=transaction,
                    produit=produit,
                    quantite=quantite,
                    prix_unitaire=produit.prix_unitaire
                )
            
            # Recalculer montant total
            transaction.calculer_montant_total()
            
            return JsonResponse({
                'success': True,
                'ligne': {
                    'id': ligne.id,
                    'produit': produit.nom,
                    'quantite': ligne.quantite,
                    'prix_unitaire': float(ligne.prix_unitaire),
                    'sous_total': float(ligne.sous_total())
                },
                'montant_total': float(transaction.montant_total),
                'montant_final': float(transaction.montant_final)
            })
            
        except Produit.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Produit introuvable.'
            })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Données JSON invalides.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur serveur: {str(e)}'
        }, status=500)


@login_required
def pos_retirer_produit(request, ligne_id):
    """
    Retire un produit de la transaction en cours
    """
    try:
        ligne = LigneTransaction.objects.get(
            id=ligne_id,
            transaction__caissier=request.user,
            transaction__statut='EN_COURS'
        )
        transaction = ligne.transaction
        ligne.delete()
        
        # Recalculer montant
        transaction.calculer_montant_total()
        
        return JsonResponse({
            'success': True,
            'montant_total': float(transaction.montant_total),
            'montant_final': float(transaction.montant_final)
        })
        
    except LigneTransaction.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Ligne introuvable.'
        })


@login_required
def pos_valider_vente(request):
    """
    Valide la vente et enregistre le(s) paiement(s)
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        paiements = data.get('paiements', [])
        client_id = data.get('client_id')
        
        # Transaction en cours
        transaction = Transaction.objects.filter(
            caissier=request.user,
            statut='EN_COURS'
        ).first()
        
        if not transaction:
            return JsonResponse({
                'success': False,
                'error': 'Aucune transaction en cours.'
            })
        
        if not transaction.lignes.exists():
            return JsonResponse({
                'success': False,
                'error': 'La transaction ne contient aucun produit.'
            })
        
        # Vérifier que le montant total des paiements correspond
        montant_paye = sum(Decimal(str(p['montant'])) for p in paiements)
        if montant_paye < transaction.montant_final:
            return JsonResponse({
                'success': False,
                'error': f'Montant insuffisant. Reçu: {montant_paye} FCFA, Requis: {transaction.montant_final} FCFA'
            })
        
        try:
            # Associer le client si fourni
            client = None
            if client_id:
                print(f"👤 ASSOCIATION CLIENT - ID: {client_id}")
                client = Client.objects.get(id=client_id)
                transaction.client = client
                print(f"   ✅ Client associé: {client.get_full_name()} (ID: {client.id})")
                # Mettre à jour dernière visite
                client.derniere_visite = timezone.now()
                client.save()
                print(f"   ✅ Dernière visite mise à jour")
            else:
                print(f"⚠️ AUCUN CLIENT_ID fourni - vente sans client")
            
            # ✅ NOUVEAU : Appliquer automatiquement le meilleur coupon
            from datetime import date
            coupon_utilise = None
            remise_coupon_montant = 0
            
            if transaction.montant_total > 0:
                aujourd_hui = date.today()
                coupons = Coupon.objects.filter(
                    statut='ACTIF',
                    date_debut__lte=aujourd_hui,
                    date_fin__gte=aujourd_hui,
                    type_coupon='GENERIC'
                ).order_by('-valeur')
                
                for coupon in coupons:
                    est_valide, message = coupon.est_valide(
                        client=client, 
                        montant_achat=float(transaction.montant_total)
                    )
                    if est_valide:
                        remise_coupon_montant = coupon.calculer_remise(float(transaction.montant_total))
                        coupon_utilise = coupon
                        
                        # Appliquer la remise à la transaction
                        transaction.montant_remise += Decimal(str(remise_coupon_montant))
                        transaction.montant_final -= Decimal(str(remise_coupon_montant))
                        transaction.save()
                        
                        print(f"🎫 COUPON APPLIQUÉ: {coupon.code} - Remise: {remise_coupon_montant} FCFA")
                        break
            
            # Enregistrer les paiements
            for paiement_data in paiements:
                type_paiement = TypePaiement.objects.get(id=paiement_data['type_id'])
                Paiement.objects.create(
                    transaction=transaction,
                    type_paiement=type_paiement,
                    montant=Decimal(str(paiement_data['montant'])),
                    reference=paiement_data.get('reference', '')
                )
            
            # Déduire le stock pour chaque ligne
            print(f"📦 MISE À JOUR DU STOCK - Transaction {transaction.numero_ticket}")
            for ligne in transaction.lignes.all():
                produit = ligne.produit
                stock_avant = produit.stock_actuel
                nouveau_stock = stock_avant - ligne.quantite
                
                # ✅ Utiliser update() au lieu de save() pour forcer la mise à jour
                Produit.objects.filter(id=produit.id).update(stock_actuel=nouveau_stock)
                
                print(f"   ✅ {produit.nom}: {stock_avant} → {nouveau_stock} ({ligne.quantite} vendus)")
                
                # Créer mouvement de stock
                MouvementStock.objects.create(
                    produit=produit,
                    type_mouvement='SORTIE',
                    quantite=-ligne.quantite,  # Négatif pour sortie
                    stock_avant=stock_avant,
                    raison=f'Vente - Ticket {transaction.numero_ticket}',
                    employe=request.user
                )
            
            print(f"📦 STOCK MIS À JOUR pour {transaction.lignes.count()} produits")
            
            # ✅ NOUVEAU : Enregistrer l'utilisation du coupon
            if coupon_utilise and client:
                UtilisationCoupon.objects.create(
                    coupon=coupon_utilise,
                    client=client,
                    transaction=transaction,
                    montant_remise=Decimal(str(remise_coupon_montant))
                )
                coupon_utilise.marquer_utilise()
                print(f"🎫 UTILISATION COUPON ENREGISTRÉE: {coupon_utilise.code}")
            
            # Valider la transaction
            transaction.statut = 'VALIDEE'
            transaction.save()
            
            # Calculer la monnaie à rendre
            monnaie = montant_paye - transaction.montant_final
            
            # Préparer les données client mises à jour si client associé
            client_data = None
            if transaction.client:
                client = transaction.client
                client_data = {
                    'id': client.id,
                    'nom_complet': client.get_full_name(),
                    'telephone': client.telephone,
                    'points': client.points_fidelite,
                    'nombre_achats': client.nombre_achats(),  # Recalculé en temps réel
                    'niveau': client.niveau_fidelite,
                    'niveau_label': client.get_niveau_fidelite_display(),
                    'remise_fidelite': get_remise_fidelite(client.niveau_fidelite)
                }
                print(f"👤 DONNÉES CLIENT MISES À JOUR: {client.get_full_name()} - {client.nombre_achats()} achats")
            
            return JsonResponse({
                'success': True,
                'numero_ticket': transaction.numero_ticket,
                'montant_final': float(transaction.montant_final),
                'montant_paye': float(montant_paye),
                'monnaie': float(monnaie),
                'client': client_data,  # ✅ NOUVEAU: Données client mises à jour
                'message': 'Vente enregistrée avec succès!'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})


@login_required
def pos_annuler_transaction(request):
    """
    Annule la transaction en cours et supprime toutes ses lignes
    """
    if request.method == 'POST':
        try:
            # Trouver la transaction EN_COURS de l'utilisateur
            transaction = Transaction.objects.filter(
                caissier=request.user,
                statut='EN_COURS'
            ).first()
            
            if transaction:
                # Supprimer toutes les lignes de la transaction
                transaction.lignes.all().delete()
                
                # Supprimer la transaction elle-même
                numero_ticket = transaction.numero_ticket
                transaction.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Transaction {numero_ticket} annulée avec succès.'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucune transaction en cours à annuler.'
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de l\'annulation: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})


@login_required
def pos_ouvrir_session(request):
    """
    Ouvre une nouvelle session de caisse
    """
    # Vérifier si une session est déjà ouverte
    session_existante = SessionCaisse.objects.filter(
        caissier=request.user,
        est_cloturee=False
    ).first()
    
    if session_existante:
        messages.warning(request, "Vous avez déjà une session ouverte.")
        return redirect('pos_interface')
    
    # Obtenir les caisses déjà occupées
    caisses_occupees = SessionCaisse.objects.filter(
        est_cloturee=False
    ).values_list('numero_caisse', flat=True)
    
    # Caisses disponibles (1 à 10)
    caisses_disponibles = [i for i in range(1, 11) if i not in caisses_occupees]
    
    if request.method == 'POST':
        fonds_ouverture = request.POST.get('fonds_ouverture')
        numero_caisse = request.POST.get('numero_caisse')
        
        if not numero_caisse:
            messages.error(request, "Veuillez sélectionner un numéro de caisse.")
            return render(request, 'caisse/pos_ouvrir_session.html', {
                'caisses_disponibles': caisses_disponibles
            })
        
        try:
            # Vérifier que la caisse est disponible
            if SessionCaisse.objects.filter(numero_caisse=numero_caisse, est_cloturee=False).exists():
                messages.error(request, f"La caisse #{numero_caisse} est déjà occupée.")
                return render(request, 'caisse/pos_ouvrir_session.html', {
                    'caisses_disponibles': caisses_disponibles
                })
            
            session = SessionCaisse.objects.create(
                caissier=request.user,
                numero_caisse=int(numero_caisse),
                fonds_ouverture=Decimal(fonds_ouverture)
            )
            messages.success(request, f"✅ Caisse #{numero_caisse} ouverte avec un fonds de {fonds_ouverture} FCFA.")
            return redirect('pos_interface')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ouverture: {str(e)}")
    
    context = {
        'caisses_disponibles': caisses_disponibles,
        'caisses_occupees': list(caisses_occupees)
    }
    
    return render(request, 'caisse/pos_ouvrir_session.html', context)


@login_required
def pos_cloturer_session(request):
    """
    Clôture la session de caisse en cours
    """
    session = SessionCaisse.objects.filter(
        caissier=request.user,
        est_cloturee=False
    ).first()
    
    if not session:
        messages.error(request, "Aucune session ouverte à clôturer.")
        return redirect('dashboard_caisse')
    
    # Calculer les statistiques de la session
    transactions = Transaction.objects.filter(
        caissier=request.user,
        date_transaction__gte=session.date_ouverture,
        statut='VALIDEE'
    )
    
    nb_transactions = transactions.count()
    total_ventes = transactions.aggregate(total=Sum('montant_final'))['total'] or 0
    
    # Répartition par type de paiement
    paiements_stats = Paiement.objects.filter(
        transaction__in=transactions
    ).values('type_paiement__nom').annotate(
        total=Sum('montant'),
        count=Count('id')
    ).order_by('-total')
    
    if request.method == 'POST':
        fonds_reel = request.POST.get('fonds_reel')
        notes = request.POST.get('notes', '')
        
        try:
            session.cloturer(Decimal(fonds_reel))
            session.notes = notes
            session.save()
            
            messages.success(request, f"Session clôturée. Écart: {session.ecart} FCFA")
            return redirect('dashboard_caisse')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la clôture: {str(e)}")
    
    context = {
        'session': session,
        'nb_transactions': nb_transactions,
        'total_ventes': total_ventes,
        'paiements_stats': paiements_stats,
        'fonds_theorique': session.calculer_fonds_theorique(),
    }
    
    return render(request, 'caisse/pos_cloturer_session.html', context)


@login_required
def dashboard_caisse(request):
    """
    Dashboard principal du module Caisse
    """
    # Session actuelle
    session_actuelle = SessionCaisse.objects.filter(
        caissier=request.user,
        est_cloturee=False
    ).first()
    
    # Statistiques du jour
    today = timezone.now().date()
    transactions_jour = Transaction.objects.filter(
        caissier=request.user,
        date_transaction__date=today,
        statut='VALIDEE'
    )
    
    nb_ventes_jour = transactions_jour.count()
    ca_jour = transactions_jour.aggregate(total=Sum('montant_final'))['total'] or 0
    
    # Dernières transactions (toutes validées du jour)
    dernieres_transactions = Transaction.objects.filter(
        caissier=request.user,
        statut='VALIDEE',
        date_transaction__date=today
    ).order_by('-date_transaction')[:10]
    
    # Transactions de la session en cours (pour historique en temps réel)
    transactions_session = []
    total_session = 0
    if session_actuelle:
        # Filtrer par caissier et date de la session au lieu de 'session'
        transactions_session = Transaction.objects.filter(
            caissier=request.user,
            statut='VALIDEE',
            date_transaction__gte=session_actuelle.date_ouverture
        ).select_related('caissier').order_by('-date_transaction')[:15]
        
        # Calculer le total de la session
        total_session = transactions_session.aggregate(total=Sum('montant_final'))['total'] or 0
    
    # Historique des sessions
    sessions = SessionCaisse.objects.filter(
        caissier=request.user
    ).order_by('-date_ouverture')[:5]
    
    context = {
        'session_actuelle': session_actuelle,
        'nb_ventes_jour': nb_ventes_jour,
        'ca_jour': ca_jour,
        'dernieres_transactions': dernieres_transactions,
        'transactions_session': transactions_session,
        'total_session': total_session,
        'sessions': sessions,
    }
    
    return render(request, 'dashboard/caisse.html', context)


# =====================================================
# SPRINT 3 - MODULE CRM & FIDÉLISATION
# =====================================================

# -------------------- Gestion Clients --------------------

def clients_list(request):
    """Liste des clients avec filtres et recherche"""
    clients = Client.objects.all().order_by('-derniere_visite')
    
    # Recherche
    search_query = request.GET.get('q', '')
    if search_query:
        clients = clients.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(numero_client__icontains=search_query) |
            Q(telephone__icontains=search_query)
        )
    
    # Filtres
    niveau = request.GET.get('niveau', '')
    if niveau:
        clients = clients.filter(niveau_fidelite=niveau)
    
    statut = request.GET.get('statut', '')
    if statut == 'actif':
        clients = clients.filter(est_actif=True)
    elif statut == 'inactif':
        clients = clients.filter(est_actif=False)
    
    # Statistiques
    stats = {
        'total': Client.objects.count(),
        'actifs': Client.objects.filter(est_actif=True).count(),
        'bronze': Client.objects.filter(niveau_fidelite='BRONZE').count(),
        'argent': Client.objects.filter(niveau_fidelite='ARGENT').count(),
        'or': Client.objects.filter(niveau_fidelite='OR').count(),
        'platine': Client.objects.filter(niveau_fidelite='PLATINE').count(),
    }
    
    context = {
        'clients': clients,
        'stats': stats,
        'search_query': search_query,
        'niveau': niveau,
        'statut': statut,
    }
    
    return render(request, 'crm/clients_list.html', context)


def client_detail(request, client_id):
    """Détail d'un client avec historique et carte fidélité"""
    client = get_object_or_404(Client, id=client_id)
    
    # Historique des achats
    ventes = Vente.objects.filter(client=client).order_by('-date_vente')[:10]
    transactions = Transaction.objects.filter(client=client).order_by('-date_transaction')[:10]
    
    # Carte de fidélité
    try:
        carte = CarteFidelite.objects.get(client=client)
        operations = OperationFidelite.objects.filter(carte=carte).order_by('-date_operation')[:20]
    except CarteFidelite.DoesNotExist:
        carte = None
        operations = []
    
    # Statistiques client
    stats = {
        'nb_achats': ventes.count() + transactions.count(),
        'montant_total': client.total_achats,
        'panier_moyen': client.total_achats / (ventes.count() + transactions.count()) if (ventes.count() + transactions.count()) > 0 else 0,
        'derniere_visite': client.derniere_visite,
        'points': client.points_fidelite,
    }
    
    context = {
        'client': client,
        'ventes': ventes,
        'transactions': transactions,
        'carte': carte,
        'operations': operations,
        'stats': stats,
    }
    
    return render(request, 'crm/client_detail.html', context)


def client_create_carte(request, client_id):
    """Créer une carte de fidélité pour un client"""
    client = get_object_or_404(Client, id=client_id)
    
    # Vérifier si le client a déjà une carte
    if hasattr(client, 'carte_fidelite'):
        messages.warning(request, "Ce client possède déjà une carte de fidélité.")
        return redirect('client_detail', client_id=client_id)
    
    if request.method == 'POST':
        # Créer la carte
        carte = CarteFidelite.objects.create(
            client=client,
            solde_points=client.points_fidelite,
            statut='ACTIVE'
        )
        
        # Créer opération initiale si points > 0
        if client.points_fidelite > 0:
            OperationFidelite.objects.create(
                carte=carte,
                type_operation='CREDIT',
                points=client.points_fidelite,
                solde_avant=0,
                solde_apres=client.points_fidelite,
                motif="Solde initial"
            )
        
        messages.success(request, f"Carte de fidélité {carte.numero_carte} créée avec succès!")
        return redirect('client_detail', client_id=client_id)
    
    context = {'client': client}
    return render(request, 'crm/client_create_carte.html', context)


def client_crediter_points(request, client_id):
    """Créditer des points à un client"""
    client = get_object_or_404(Client, id=client_id)
    
    try:
        carte = CarteFidelite.objects.get(client=client)
    except CarteFidelite.DoesNotExist:
        messages.error(request, "Ce client n'a pas de carte de fidélité.")
        return redirect('client_detail', client_id=client_id)
    
    if request.method == 'POST':
        points = int(request.POST.get('points', 0))
        motif = request.POST.get('motif', '')
        
        if points > 0:
            carte.crediter_points(points, motif)
            client.points_fidelite += points
            client.save()
            
            messages.success(request, f"{points} points crédités avec succès!")
        else:
            messages.error(request, "Le nombre de points doit être positif.")
        
        return redirect('client_detail', client_id=client_id)
    
    context = {'client': client, 'carte': carte}
    return render(request, 'crm/client_crediter_points.html', context)


# -------------------- Segments Clients --------------------

def segments_list(request):
    """Liste des segments clients"""
    segments = SegmentClient.objects.all().order_by('nom')
    
    # Ajouter le nombre de clients pour chaque segment
    segments_with_count = []
    for segment in segments:
        segments_with_count.append({
            'segment': segment,
            'nb_clients': segment.nombre_clients()
        })
    
    context = {'segments': segments_with_count}
    return render(request, 'crm/segments_list.html', context)


def segment_create(request):
    """Créer un nouveau segment client"""
    if request.method == 'POST':
        nom = request.POST.get('nom')
        description = request.POST.get('description', '')
        niveau_fidelite = request.POST.get('niveau_fidelite', '')
        montant_min = request.POST.get('montant_achats_min', None)
        montant_max = request.POST.get('montant_achats_max', None)
        date_visite = request.POST.get('date_derniere_visite_avant', None)
        
        segment = SegmentClient.objects.create(
            nom=nom,
            description=description,
            niveau_fidelite=niveau_fidelite if niveau_fidelite else None,
            montant_achats_min=Decimal(montant_min) if montant_min else None,
            montant_achats_max=Decimal(montant_max) if montant_max else None,
            date_derniere_visite_avant=date_visite if date_visite else None
        )
        
        messages.success(request, f"Segment '{segment.nom}' créé avec succès!")
        return redirect('segments_list')
    
    context = {
        'niveaux_fidelite': Client.NIVEAUX_FIDELITE
    }
    return render(request, 'crm/segment_create.html', context)


def segment_clients(request, segment_id):
    """Afficher les clients d'un segment"""
    segment = get_object_or_404(SegmentClient, id=segment_id)
    clients = segment.get_clients()
    
    context = {
        'segment': segment,
        'clients': clients,
        'nb_clients': clients.count()
    }
    
    return render(request, 'crm/segment_clients.html', context)


# -------------------- Campagnes Marketing --------------------

def campagnes_list(request):
    """Liste des campagnes marketing"""
    campagnes = Campagne.objects.all().order_by('-date_creation')
    
    # Statistiques
    stats = {
        'total': campagnes.count(),
        'en_cours': campagnes.filter(statut='EN_COURS').count(),
        'programmees': campagnes.filter(statut='PROGRAMMEE').count(),
        'terminees': campagnes.filter(statut='TERMINEE').count(),
    }
    
    context = {
        'campagnes': campagnes,
        'stats': stats
    }
    
    return render(request, 'crm/campagnes_list.html', context)


def campagne_create(request):
    """Créer une nouvelle campagne"""
    if request.method == 'POST':
        titre = request.POST.get('titre')
        description = request.POST.get('description', '')
        type_campagne = request.POST.get('type_campagne')
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        segment_id = request.POST.get('segment_cible')
        message = request.POST.get('message')
        
        segment = SegmentClient.objects.get(id=segment_id) if segment_id else None
        nb_destinataires = segment.nombre_clients() if segment else Client.objects.filter(est_actif=True).count()
        
        campagne = Campagne.objects.create(
            titre=titre,
            description=description,
            type_campagne=type_campagne,
            date_debut=date_debut,
            date_fin=date_fin,
            segment_cible=segment,
            message=message,
            nb_destinataires=nb_destinataires,
            creee_par=request.user.employe if hasattr(request.user, 'employe') else None,
            statut='PROGRAMMEE'
        )
        
        messages.success(request, f"Campagne '{campagne.titre}' créée avec succès!")
        return redirect('campagnes_list')
    
    segments = SegmentClient.objects.filter(est_actif=True)
    
    context = {
        'segments': segments,
        'types': Campagne.TYPE_CHOICES
    }
    
    return render(request, 'crm/campagne_create.html', context)


def campagne_detail(request, campagne_id):
    """Détail d'une campagne"""
    campagne = get_object_or_404(Campagne, id=campagne_id)
    
    # Clients ciblés
    if campagne.segment_cible:
        clients = campagne.segment_cible.get_clients()
    else:
        clients = Client.objects.filter(est_actif=True)
    
    # Statistiques
    stats = {
        'destinataires': campagne.nb_destinataires,
        'envoyes': campagne.nb_envoyes,
        'ouverts': campagne.nb_ouverts,
        'taux_ouverture': campagne.taux_ouverture(),
        'taux_envoi': (campagne.nb_envoyes / campagne.nb_destinataires * 100) if campagne.nb_destinataires > 0 else 0
    }
    
    context = {
        'campagne': campagne,
        'clients': clients[:20],  # Limiter l'affichage
        'stats': stats
    }
    
    return render(request, 'crm/campagne_detail.html', context)


def campagne_send(request, campagne_id):
    """Lancer l'envoi d'une campagne"""
    campagne = get_object_or_404(Campagne, id=campagne_id)
    
    if request.method == 'POST':
        # Obtenir les clients ciblés
        if campagne.segment_cible:
            clients = campagne.segment_cible.get_clients()
        else:
            clients = Client.objects.filter(est_actif=True)
        
        # Simulation d'envoi (dans la vraie vie, on utiliserait une API SMS/Email)
        envois_reussis = 0
        for client in clients:
            # Ici on simule l'envoi
            # Dans la réalité: envoyer_sms(client.telephone, campagne.message)
            # ou envoyer_email(client.email, campagne.message)
            envois_reussis += 1
        
        # Mettre à jour la campagne
        campagne.nb_envoyes = envois_reussis
        campagne.statut = 'EN_COURS'
        campagne.save()
        
        messages.success(request, f"Campagne lancée! {envois_reussis} messages envoyés.")
        return redirect('campagne_detail', campagne_id=campagne_id)
    
    context = {'campagne': campagne}
    return render(request, 'crm/campagne_send.html', context)


# =====================================================
# SPRINT 4 - MODULE ANALYTICS & REPORTING
# =====================================================

def dashboard_analytics(request):
    """Dashboard analytique complet avec graphiques"""
    from datetime import timedelta
    
    # Période d'analyse
    today = timezone.now().date()
    debut_mois = today.replace(day=1)
    debut_semaine = today - timedelta(days=today.weekday())
    
    # === Ventes & CA ===
    ventes_mois = Transaction.objects.filter(
        date_transaction__gte=debut_mois,
        statut='VALIDEE'
    )
    
    ca_mois = ventes_mois.aggregate(total=Sum('montant_final'))['total'] or 0
    nb_ventes_mois = ventes_mois.count()
    panier_moyen_mois = ca_mois / nb_ventes_mois if nb_ventes_mois > 0 else 0
    
    # CA par jour (30 derniers jours)
    ca_par_jour = []
    for i in range(30):
        jour = today - timedelta(days=29-i)
        ca_jour = Transaction.objects.filter(
            date_transaction__date=jour,
            statut='VALIDEE'
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        ca_par_jour.append({
            'date': jour.strftime('%d/%m'),
            'montant': float(ca_jour)
        })
    
    # === Produits ===
    # Top 10 produits vendus
    from django.db.models import Count
    top_produits = LigneTransaction.objects.filter(
        transaction__statut='VALIDEE',
        transaction__date_transaction__gte=debut_mois
    ).values(
        'produit__nom'
    ).annotate(
        quantite=Sum('quantite'),
        ca=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-quantite')[:10]
    
    # === Clients ===
    clients_actifs = Client.objects.filter(
        derniere_visite__gte=debut_mois
    ).count()
    
    nouveaux_clients = Client.objects.filter(
        date_inscription__gte=debut_mois
    ).count() if hasattr(Client, 'date_inscription') else 0
    
    # Répartition par niveau
    repartition_niveaux = []
    for niveau, label in Client.NIVEAUX_FIDELITE:
        nb = Client.objects.filter(niveau_fidelite=niveau).count()
        repartition_niveaux.append({'niveau': label, 'nombre': nb})
    
    # === Stock ===
    valeur_stock = Produit.objects.aggregate(
        total=Sum(F('stock_actuel') * F('prix_achat'))
    )['total'] or 0
    
    produits_alerte = Produit.objects.filter(
        stock_actuel__lte=F('stock_minimum')
    ).count()
    
    # === Tendances hebdomadaires ===
    ca_hebdo = []
    for i in range(7):
        jour = debut_semaine + timedelta(days=i)
        if jour > today:
            break
        ca = Transaction.objects.filter(
            date_transaction__date=jour,
            statut='VALIDEE'
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        ca_hebdo.append({
            'jour': jour.strftime('%A'),
            'montant': float(ca)
        })
    
    # === Performances caissiers ===
    top_caissiers = Transaction.objects.filter(
        date_transaction__gte=debut_mois,
        statut='VALIDEE'
    ).values(
        'caissier__first_name', 'caissier__last_name'
    ).annotate(
        nb_ventes=Count('id'),
        ca_total=Sum('montant_final')
    ).order_by('-ca_total')[:5]
    
    context = {
        # KPIs principaux
        'ca_mois': ca_mois,
        'nb_ventes_mois': nb_ventes_mois,
        'panier_moyen_mois': panier_moyen_mois,
        'clients_actifs': clients_actifs,
        'nouveaux_clients': nouveaux_clients,
        'valeur_stock': valeur_stock,
        'produits_alerte': produits_alerte,
        
        # Graphiques (JSON)
        'ca_par_jour': json.dumps(ca_par_jour),
        'ca_hebdo': json.dumps(ca_hebdo),
        'top_produits': list(top_produits),
        'repartition_niveaux': json.dumps(repartition_niveaux),
        'top_caissiers': list(top_caissiers),
    }
    
    return render(request, 'analytics/dashboard_analytics.html', context)


def export_ventes_excel(request):
    """Exporter les ventes en Excel"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Récupérer les paramètres de filtrage
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Requête de base
    transactions = Transaction.objects.filter(statut='VALIDEE').order_by('-date_transaction')
    
    if date_debut:
        transactions = transactions.filter(date_transaction__gte=date_debut)
    if date_fin:
        transactions = transactions.filter(date_transaction__lte=date_fin)
    
    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"
    
    # En-têtes
    headers = ['Date', 'N° Ticket', 'Caissier', 'Client', 'Montant Total', 'Remise', 'Montant Final', 'Statut']
    ws.append(headers)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for transaction in transactions:
        ws.append([
            transaction.date_transaction.strftime('%d/%m/%Y %H:%M'),
            transaction.numero_ticket,
            transaction.caissier.get_full_name() if transaction.caissier else '—',
            transaction.client.get_full_name() if transaction.client else '—',
            float(transaction.montant_total),
            float(transaction.montant_remise),  # ✅ Corrigé
            float(transaction.montant_final),
            transaction.get_statut_display()
        ])
    
    # Ajuster largeurs
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Préparer la réponse HTTP
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from datetime import date
    today = date.today()
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ventes_{today.strftime("%Y%m%d")}.xlsx'
    
    return response


def rapport_mensuel(request):
    """Génération de rapport mensuel complet"""
    from datetime import timedelta
    from calendar import monthrange
    
    # Paramètres
    annee = int(request.GET.get('annee', timezone.now().year))
    mois = int(request.GET.get('mois', timezone.now().month))
    
    # Dates
    premier_jour = timezone.datetime(annee, mois, 1).date()
    dernier_jour = timezone.datetime(annee, mois, monthrange(annee, mois)[1]).date()
    
    # Transactions du mois
    transactions = Transaction.objects.filter(
        date_transaction__date__gte=premier_jour,
        date_transaction__date__lte=dernier_jour,
        statut='VALIDEE'
    )
    
    # === Calculs ===
    # CA
    ca_total = transactions.aggregate(total=Sum('montant_final'))['total'] or 0
    nb_ventes = transactions.count()
    panier_moyen = ca_total / nb_ventes if nb_ventes > 0 else 0
    
    # Clients
    nb_clients_uniques = transactions.values('client').distinct().count()
    nouveaux_clients = Client.objects.filter(
        date_inscription__gte=premier_jour,
        date_inscription__lte=dernier_jour
    ).count() if hasattr(Client, 'date_inscription') else 0
    
    # Produits
    produits_vendus = LigneTransaction.objects.filter(
        transaction__in=transactions
    ).aggregate(total=Sum('quantite'))['total'] or 0
    
    # Top 5 produits
    top_produits = LigneTransaction.objects.filter(
        transaction__in=transactions
    ).values(
        'produit__nom'
    ).annotate(
        quantite=Sum('quantite'),
        ca=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-ca')[:5]
    
    # Ventes par jour
    ventes_par_jour = []
    for i in range((dernier_jour - premier_jour).days + 1):
        jour = premier_jour + timedelta(days=i)
        ca_jour = Transaction.objects.filter(
            date_transaction__date=jour,
            statut='VALIDEE'
        ).aggregate(total=Sum('montant_final'))['total'] or 0
        nb_jour = Transaction.objects.filter(
            date_transaction__date=jour,
            statut='VALIDEE'
        ).count()
        ventes_par_jour.append({
            'date': jour,
            'ca': ca_jour,
            'nb_ventes': nb_jour
        })
    
    # Performances caissiers
    caissiers = transactions.values(
        'caissier__first_name', 'caissier__last_name'
    ).annotate(
        nb_ventes=Count('id'),
        ca_total=Sum('montant_final')
    ).order_by('-ca_total')
    
    context = {
        'annee': annee,
        'mois': mois,
        'nom_mois': timezone.datetime(annee, mois, 1).strftime('%B'),
        'ca_total': ca_total,
        'nb_ventes': nb_ventes,
        'panier_moyen': panier_moyen,
        'nb_clients_uniques': nb_clients_uniques,
        'nouveaux_clients': nouveaux_clients,
        'produits_vendus': produits_vendus,
        'top_produits': list(top_produits),
        'ventes_par_jour': ventes_par_jour,
        'caissiers': list(caissiers),
    }
    
    return render(request, 'analytics/rapport_mensuel.html', context)


# ==================== GESTION COMMANDES FOURNISSEURS (Scénario 8.1.1) ====================

@login_required
def commandes_fournisseurs(request):
    """Liste des commandes fournisseurs avec filtres"""
    if request.user.role not in ['STOCK', 'ADMIN', 'MANAGER']:
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    # Filtres
    statut_filtre = request.GET.get('statut', '')
    fournisseur_filtre = request.GET.get('fournisseur', '')
    
    commandes = CommandeFournisseur.objects.all().select_related('fournisseur', 'employe')
    
    if statut_filtre:
        commandes = commandes.filter(statut=statut_filtre)
    if fournisseur_filtre:
        commandes = commandes.filter(fournisseur_id=fournisseur_filtre)
    
    commandes = commandes.order_by('-date_commande')[:50]
    
    # Statistiques
    total_commandes = CommandeFournisseur.objects.count()
    en_attente = CommandeFournisseur.objects.filter(statut='EN_ATTENTE').count()
    validees = CommandeFournisseur.objects.filter(statut='VALIDEE').count()
    livrees = CommandeFournisseur.objects.filter(statut='LIVREE').count()
    
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    
    context = {
        'commandes': commandes,
        'fournisseurs': fournisseurs,
        'total_commandes': total_commandes,
        'en_attente': en_attente,
        'validees': validees,
        'livrees': livrees,
        'statut_filtre': statut_filtre,
        'fournisseur_filtre': fournisseur_filtre,
    }
    
    return render(request, 'dashboard/commandes_fournisseurs.html', context)


@login_required
def creer_commande_fournisseur(request):
    """Créer une commande fournisseur (Scénario 8.1.1)"""
    if request.user.role not in ['STOCK', 'ADMIN', 'MANAGER']:
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        fournisseur_id = request.POST.get('fournisseur')
        produit_id = request.POST.get('produit')
        quantite = int(request.POST.get('quantite', 0))
        
        try:
            fournisseur = Fournisseur.objects.get(id=fournisseur_id)
            produit = Produit.objects.get(id=produit_id)
            
            # Générer numéro de commande
            dernier_numero = CommandeFournisseur.objects.count() + 1
            numero_commande = f"CF{timezone.now().strftime('%Y%m%d')}{dernier_numero:04d}"
            
            # Calculer montant total
            montant_total = produit.prix_achat * quantite
            
            # Créer la commande
            commande = CommandeFournisseur.objects.create(
                numero_commande=numero_commande,
                fournisseur=fournisseur,
                employe=request.user.employe if hasattr(request.user, 'employe') else None,
                statut='EN_ATTENTE',
                date_commande=timezone.now(),
                date_livraison_prevue=timezone.now() + timedelta(days=fournisseur.delai_livraison_moyen),
                montant_total=montant_total
            )
            
            # Ajouter la ligne
            LigneCommandeFournisseur.objects.create(
                commande=commande,
                produit=produit,
                quantite=quantite,
                prix_unitaire=produit.prix_achat
            )
            
            messages.success(request, f'✅ Commande créée! {quantite} unités de {produit.nom} commandées auprès de {fournisseur.nom}. Livraison prévue le {commande.date_livraison_prevue.strftime("%d/%m/%Y")}.')
            return redirect('commandes_fournisseurs')
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    # GET: afficher formulaire
    produits_critiques = []
    for produit in Produit.objects.filter(est_actif=True):
        if produit.besoin_reapprovisionnement():
            produits_critiques.append({
                'produit': produit,
                'quantite_recommandee': produit.quantite_a_commander(),
                'fournisseur': produit.fournisseur_principal
            })
    
    fournisseurs = Fournisseur.objects.filter(est_actif=True)
    produits = Produit.objects.filter(est_actif=True).order_by('nom')
    
    context = {
        'produits_critiques': produits_critiques,
        'fournisseurs': fournisseurs,
        'produits': produits,
    }
    
    return render(request, 'dashboard/creer_commande_fournisseur.html', context)


@login_required
def valider_commande_fournisseur(request, commande_id):
    """Valider une commande fournisseur"""
    if request.user.role not in ['STOCK', 'ADMIN', 'MANAGER']:
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    try:
        commande = CommandeFournisseur.objects.get(id=commande_id)
        commande.statut = 'VALIDEE'
        commande.save()
        
        messages.success(request, f'✅ Commande #{commande.id} validée! Envoyée au fournisseur {commande.fournisseur.nom}.')
    except CommandeFournisseur.DoesNotExist:
        messages.error(request, 'Commande introuvable.')
    
    return redirect('commandes_fournisseurs')


@login_required
def recevoir_commande_fournisseur(request, commande_id):
    """Marquer commande comme livrée et mettre à jour stocks"""
    if request.user.role not in ['STOCK', 'ADMIN', 'MANAGER']:
        messages.error(request, "Accès refusé.")
        return redirect('dashboard')
    
    try:
        commande = CommandeFournisseur.objects.get(id=commande_id)
        
        if commande.statut != 'VALIDEE':
            messages.warning(request, 'Cette commande doit d\'abord être validée.')
            return redirect('commandes_fournisseurs')
        
        # Mettre à jour stocks pour chaque ligne
        for ligne in commande.lignes.all():
            stock_avant = ligne.produit.stock_actuel
            ligne.produit.stock_actuel += ligne.quantite
            stock_apres = ligne.produit.stock_actuel
            ligne.produit.save()
            
            # Enregistrer mouvement
            MouvementStock.objects.create(
                produit=ligne.produit,
                type_mouvement='ENTREE',
                quantite=ligne.quantite,
                raison=f'Réception commande {commande.numero_commande}',
                employe=request.user.employe if hasattr(request.user, 'employe') else None,
                stock_avant=stock_avant,
                stock_apres=stock_apres,
                commande_fournisseur=commande
            )
        
        # Marquer commande comme livrée
        commande.statut = 'LIVREE'
        commande.date_livraison_reelle = timezone.now().date()
        commande.save()
        
        # Résoudre alertes stock
        for ligne in commande.lignes.all():
            AlerteStock.objects.filter(
                produit=ligne.produit,
                type_alerte='SEUIL_CRITIQUE',
                est_resolue=False
            ).update(est_resolue=True)
        
        messages.success(request, f'✅ Commande #{commande.id} reçue! Stocks mis à jour automatiquement.')
        
    except CommandeFournisseur.DoesNotExist:
        messages.error(request, 'Commande introuvable.')
    
    return redirect('commandes_fournisseurs')


# ==================== GESTION CAISSE (Scénario 8.1.2) ====================

@login_required
def caisse_vente(request):
    """Interface de caisse pour ventes (Scénario 8.1.2)"""
    if request.user.role not in ['CAISSIER', 'ADMIN', 'MANAGER']:
        messages.error(request, "Accès refusé. Réservé aux caissiers.")
        return redirect('dashboard')
    
    # Initialiser le panier en session si n'existe pas
    if 'panier' not in request.session:
        request.session['panier'] = []
        request.session['client_id'] = None
    
    # Récupérer données
    produits = Produit.objects.filter(est_actif=True, stock_actuel__gt=0).order_by('nom')
    clients = Client.objects.all().order_by('nom')
    
    # Récupérer panier de la session
    panier = request.session.get('panier', [])
    client_id = request.session.get('client_id')
    client = None
    if client_id:
        try:
            client = Client.objects.get(id=client_id)
        except Client.DoesNotExist:
            pass
    
    # Calculer stock disponible (stock réel - quantités dans panier)
    produits_avec_stock = []
    for produit in produits:
        quantite_panier = sum(item['quantite'] for item in panier if item['produit_id'] == str(produit.id))
        stock_disponible = produit.stock_actuel - quantite_panier
        produit.stock_disponible = stock_disponible  # Ajouter attribut temporaire
        produits_avec_stock.append(produit)
    
    # Calculer totaux
    sous_total = sum(item['montant_ligne'] for item in panier)
    
    # Calcul remises
    remise_fidelite = 0
    remise_promotionnelle = 0
    pourcentage_fidelite = 0
    
    if client:
        if client.niveau_fidelite == 'VIP':
            pourcentage_fidelite = 10
            remise_fidelite = sous_total * 0.10
        elif client.niveau_fidelite == 'GOLD':
            pourcentage_fidelite = 5
            remise_fidelite = sous_total * 0.05
        elif client.niveau_fidelite == 'SILVER':
            pourcentage_fidelite = 3
            remise_fidelite = sous_total * 0.03
    
    # Remise promotionnelle (5% si ≥ 40,000 FCFA)
    montant_apres_fidelite = sous_total - remise_fidelite
    if montant_apres_fidelite >= 40000:
        remise_promotionnelle = montant_apres_fidelite * 0.05
    
    # ✅ NOUVEAU : Recherche et application automatique des coupons
    from datetime import date
    from decimal import Decimal
    
    coupons_disponibles = []
    coupon_applique = None
    remise_coupon = 0
    
    if sous_total > 0:  # Seulement si le panier n'est pas vide
        aujourd_hui = date.today()
        
        # Chercher les coupons actifs et valides
        coupons = Coupon.objects.filter(
            statut='ACTIF',
            date_debut__lte=aujourd_hui,
            date_fin__gte=aujourd_hui
        ).order_by('-valeur')  # Prioriser les plus avantageux
        
        for coupon in coupons:
            # Vérifier si le coupon est applicable
            est_valide, message = coupon.est_valide(client=client, montant_achat=sous_total)
            
            if est_valide:
                # Calculer la remise du coupon
                remise_calculee = coupon.calculer_remise(sous_total)
                
                # Ajouter à la liste des coupons disponibles
                coupons_disponibles.append({
                    'coupon': coupon,
                    'remise': remise_calculee,
                    'est_applique': False
                })
                
                # Appliquer automatiquement le meilleur coupon GENERIC
                # Pour les coupons SPECIAL, on les affiche mais on attend confirmation
                if coupon.type_coupon == 'GENERIC' and coupon_applique is None:
                    coupon_applique = coupon
                    remise_coupon = remise_calculee
                    coupons_disponibles[-1]['est_applique'] = True
    
    total_remises = remise_fidelite + remise_promotionnelle + remise_coupon
    montant_avant_tva = sous_total - total_remises
    tva = montant_avant_tva * 0.18  # TVA 18%
    montant_final = montant_avant_tva + tva
    
    # Points fidélité à gagner
    points_a_gagner = int(montant_final / 1000) if montant_final > 0 else 0
    
    context = {
        'produits': produits_avec_stock,  # Utiliser la liste avec stock disponible
        'clients': clients,
        'panier': panier,
        'client': client,
        'sous_total': sous_total,
        'remise_fidelite': remise_fidelite,
        'pourcentage_fidelite': pourcentage_fidelite,
        'remise_promotionnelle': remise_promotionnelle,
        'remise_coupon': remise_coupon,  # ✅ NOUVEAU
        'coupon_applique': coupon_applique,  # ✅ NOUVEAU
        'coupons_disponibles': coupons_disponibles,  # ✅ NOUVEAU
        'total_remises': total_remises,
        'tva': tva,
        'montant_final': montant_final,
        'points_a_gagner': points_a_gagner,
        'nb_articles': len(panier),
        'caisse_numero': request.user.employe.caisse_assignee if hasattr(request.user, 'employe') and hasattr(request.user.employe, 'caisse_assignee') else '#01',
    }
    
    return render(request, 'caisse/index.html', context)


@login_required
def caisse_ajouter_produit(request):
    """AJAX - Ajouter un produit au panier"""
    if request.method == 'POST':
        produit_id = request.POST.get('produit_id')
        quantite = int(request.POST.get('quantite', 1))
        
        try:
            produit = Produit.objects.get(id=produit_id, est_actif=True)
            
            # Vérifier stock
            if produit.stock_actuel < quantite:
                return JsonResponse({'success': False, 'error': f'Stock insuffisant. Disponible: {produit.stock_actuel}'})
            
            # Récupérer panier
            panier = request.session.get('panier', [])
            
            # Vérifier si produit déjà dans panier
            produit_existe = False
            for item in panier:
                if item['produit_id'] == produit_id:
                    item['quantite'] += quantite
                    item['montant_ligne'] = item['quantite'] * item['prix_unitaire']
                    produit_existe = True
                    break
            
            # Si pas dans panier, ajouter
            if not produit_existe:
                panier.append({
                    'produit_id': produit_id,
                    'nom': produit.nom,
                    'quantite': quantite,
                    'prix_unitaire': float(produit.prix_unitaire),
                    'montant_ligne': float(produit.prix_unitaire * quantite)
                })
            
            request.session['panier'] = panier
            request.session.modified = True
            
            # Calculer la quantité totale dans le panier pour ce produit
            quantite_panier = sum(item['quantite'] for item in panier if item['produit_id'] == produit_id)
            stock_disponible = produit.stock_actuel - quantite_panier
            
            return JsonResponse({
                'success': True, 
                'message': f'{produit.nom} ajouté au panier',
                'produit_id': produit_id,
                'stock_disponible': stock_disponible,
                'quantite_panier': quantite_panier
            })
            
        except Produit.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Produit introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def caisse_retirer_produit(request, produit_id):
    """Retirer un produit du panier"""
    panier = request.session.get('panier', [])
    panier = [item for item in panier if item['produit_id'] != produit_id]
    request.session['panier'] = panier
    request.session.modified = True
    
    messages.success(request, 'Produit retiré du panier')
    return redirect('caisse_vente')


@login_required
def caisse_vider_panier(request):
    """Vider le panier"""
    request.session['panier'] = []
    request.session['client_id'] = None
    request.session.modified = True
    
    messages.success(request, 'Panier vidé')
    return redirect('caisse_vente')


@login_required
def caisse_identifier_client(request):
    """AJAX - Identifier ou créer un client par téléphone"""
    if request.method == 'POST':
        import json
        telephone = request.POST.get('telephone', '').strip()
        action = request.POST.get('action', 'identify')  # identify ou create
        
        if not telephone:
            request.session['client_id'] = None
            request.session.modified = True
            return JsonResponse({'success': True, 'client': None, 'message': 'Vente anonyme'})
        
        # Vérifier si le client existe
        try:
            client = Client.objects.get(telephone=telephone)
            client.derniere_visite = timezone.now()
            client.save()
            
            request.session['client_id'] = client.id
            request.session.modified = True
            
            # Récupérer les coupons disponibles (si modèle Coupon existe)
            coupons_disponibles = []
            try:
                from datetime import date
                coupons = Coupon.objects.filter(
                    client=client,
                    est_utilise=False,
                    date_debut__lte=date.today(),
                    date_fin__gte=date.today()
                )
                for coupon in coupons:
                    coupons_disponibles.append({
                        'id': coupon.id,
                        'code': coupon.code,
                        'type': coupon.type_remise,
                        'valeur': float(coupon.valeur),
                        'description': coupon.description
                    })
            except:
                pass
            
            return JsonResponse({
                'success': True,
                'action': 'identified',
                'client': {
                    'id': client.id,
                    'numero': client.numero_client,
                    'nom_complet': client.get_full_name(),
                    'nom': client.nom,
                    'prenom': client.prenom,
                    'telephone': client.telephone,
                    'email': client.email,
                    'niveau': client.niveau_fidelite,
                    'niveau_label': client.get_niveau_fidelite_display(),
                    'points': client.points_fidelite,
                    'total_achats': float(client.total_achats),
                    'nombre_achats': client.nombre_achats(),
                    'coupons': coupons_disponibles,
                    'remise_fidelite': get_remise_fidelite(client.niveau_fidelite)
                },
                'message': f'Client {client.get_full_name()} identifié!'
            })
            
        except Client.DoesNotExist:
            # Si action = create, créer le nouveau client
            if action == 'create':
                nom = request.POST.get('nom', '').strip()
                prenom = request.POST.get('prenom', '').strip()
                email = request.POST.get('email', '').strip()
                
                if not nom:
                    return JsonResponse({
                        'success': False,
                        'error': 'Le nom est obligatoire pour créer un nouveau client'
                    })
                
                # Créer le nouveau client
                client = Client.objects.create(
                    telephone=telephone,
                    nom=nom,
                    prenom=prenom,
                    email=email,
                    points_fidelite=0,
                    niveau_fidelite='TOUS',
                    derniere_visite=timezone.now()
                )
                
                request.session['client_id'] = client.id
                request.session.modified = True
                
                return JsonResponse({
                    'success': True,
                    'action': 'created',
                    'client': {
                        'id': client.id,
                        'numero': client.numero_client,
                        'nom_complet': client.get_full_name(),
                        'nom': client.nom,
                        'prenom': client.prenom,
                        'telephone': client.telephone,
                        'email': client.email,
                        'niveau': client.niveau_fidelite,
                        'niveau_label': client.get_niveau_fidelite_display(),
                        'points': client.points_fidelite,
                        'total_achats': 0,
                        'nombre_achats': 0,
                        'coupons': [],
                        'remise_fidelite': 0
                    },
                    'message': f'Nouveau client {client.get_full_name()} créé avec succès!'
                })
            else:
                # Client non trouvé, demander confirmation pour création
                return JsonResponse({
                    'success': False,
                    'action': 'not_found',
                    'error': 'Client non trouvé',
                    'message': f'Aucun client avec le numéro {telephone}. Voulez-vous créer un nouveau client?',
                    'telephone': telephone
                })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


def get_remise_fidelite(niveau):
    """Retourne le pourcentage de remise selon le niveau de fidélité"""
    remises = {
        'TOUS': 0,
        'SILVER': 3,
        'GOLD': 5,
        'VIP': 10
    }
    return remises.get(niveau, 0)


@login_required
def caisse_valider_vente(request):
    """Finaliser et enregistrer la vente"""
    if request.method == 'POST':
        moyen_paiement = request.POST.get('moyen_paiement')
        
        # Récupérer panier
        panier = request.session.get('panier', [])
        if not panier:
            messages.error(request, 'Panier vide!')
            return redirect('caisse_vente')
        
        # Récupérer client
        client_id = request.session.get('client_id')
        client = None
        if client_id:
            try:
                client = Client.objects.get(id=client_id)
            except Client.DoesNotExist:
                pass
        
        try:
            # Calculer totaux
            sous_total = sum(item['montant_ligne'] for item in panier)
            
            remise_fidelite = 0
            if client:
                if client.niveau_fidelite == 'VIP':
                    remise_fidelite = sous_total * 0.10
                elif client.niveau_fidelite == 'GOLD':
                    remise_fidelite = sous_total * 0.05
                elif client.niveau_fidelite == 'SILVER':
                    remise_fidelite = sous_total * 0.03
            
            montant_apres_fidelite = sous_total - remise_fidelite
            remise_promotionnelle = 0
            if montant_apres_fidelite >= 40000:
                remise_promotionnelle = montant_apres_fidelite * 0.05
            
            # ✅ NOUVEAU : Appliquer automatiquement le meilleur coupon
            from datetime import date
            from decimal import Decimal
            
            coupon_utilise = None
            remise_coupon = 0
            
            if sous_total > 0:
                aujourd_hui = date.today()
                coupons = Coupon.objects.filter(
                    statut='ACTIF',
                    date_debut__lte=aujourd_hui,
                    date_fin__gte=aujourd_hui,
                    type_coupon='GENERIC'  # Automatique pour les génériques
                ).order_by('-valeur')
                
                for coupon in coupons:
                    est_valide, message = coupon.est_valide(client=client, montant_achat=sous_total)
                    if est_valide:
                        remise_coupon = coupon.calculer_remise(sous_total)
                        coupon_utilise = coupon
                        break  # Prendre le premier (meilleur)
            
            total_remises = remise_fidelite + remise_promotionnelle + remise_coupon
            montant_avant_tva = sous_total - total_remises
            tva = montant_avant_tva * 0.18
            montant_final = montant_avant_tva + tva
            
            # Créer la vente
            vente = Vente.objects.create(
                caissier=request.user.employe if hasattr(request.user, 'employe') else None,
                client=client,
                montant_total=sous_total,
                montant_tva=tva,
                remise=total_remises,
                montant_final=montant_final,
                moyen_paiement=moyen_paiement,
                caisse_numero=request.user.employe.caisse_assignee if hasattr(request.user, 'employe') and hasattr(request.user.employe, 'caisse_assignee') else '#01'
            )
            
            # Créer lignes de vente et mettre à jour stock
            for item in panier:
                produit = Produit.objects.get(id=item['produit_id'])
                
                # Créer ligne
                LigneVente.objects.create(
                    vente=vente,
                    produit=produit,
                    quantite=item['quantite'],
                    prix_unitaire=item['prix_unitaire'],
                    montant_ligne=item['montant_ligne']
                )
                
                # Mettre à jour stock
                stock_avant = produit.stock_actuel
                produit.stock_actuel -= item['quantite']
                produit.save()
                
                # Créer mouvement stock
                MouvementStock.objects.create(
                    produit=produit,
                    type_mouvement='SORTIE',
                    quantite=-item['quantite'],
                    raison=f'Vente {vente.numero_transaction}',
                    employe=request.user.employe if hasattr(request.user, 'employe') else None,
                    stock_avant=stock_avant,
                    stock_apres=produit.stock_actuel
                )
            
            # Attribuer points fidélité
            if client:
                points_gagnes = int(montant_final / 1000)
                client.points_fidelite += points_gagnes
                client.save()
            
            # ✅ NOUVEAU : Enregistrer l'utilisation du coupon
            if coupon_utilise:
                UtilisationCoupon.objects.create(
                    coupon=coupon_utilise,
                    client=client,
                    vente=vente,
                    montant_remise=remise_coupon
                )
                coupon_utilise.marquer_utilise()
                print(f"🎫 COUPON UTILISÉ: {coupon_utilise.code} - Remise: {remise_coupon} FCFA")
            
            # Vider panier
            request.session['panier'] = []
            request.session['client_id'] = None
            request.session.modified = True
            
            messages.success(request, f'✅ Vente {vente.numero_transaction} enregistrée! Montant: {montant_final:,.0f} FCFA')
            return redirect('caisse_vente')
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
            return redirect('caisse_vente')
    
    return redirect('caisse_vente')


@login_required
def caisse_rapport_journalier(request):
    """Rapport des ventes du jour (Scénario 8.1.2)"""
    if request.user.role not in ['CAISSIER', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    # Date du jour
    aujourd_hui = timezone.now().date()
    date_param = request.GET.get('date')
    if date_param:
        try:
            aujourd_hui = datetime.strptime(date_param, '%Y-%m-%d').date()
        except:
            pass
    
    # Ventes du jour (UTILISER Transaction au lieu de Vente)
    ventes = Transaction.objects.filter(
        date_transaction__date=aujourd_hui,
        statut='VALIDEE'  # ✅ IMPORTANT: Seulement les ventes validées
    )
    
    # Statistiques générales
    nb_transactions = ventes.count()
    ca_total = ventes.aggregate(total=Sum('montant_final'))['total'] or 0
    remises_total = ventes.aggregate(total=Sum('montant_remise'))['total'] or 0
    panier_moyen = ca_total / nb_transactions if nb_transactions > 0 else 0
    
    # Répartition moyens de paiement
    moyens_paiement = Paiement.objects.filter(
        transaction__date_transaction__date=aujourd_hui,
        transaction__statut='VALIDEE'
    ).values('type_paiement__nom').annotate(  # ✅ Changé code → nom
        total=Sum('montant'),
        nb=Count('id')
    ).order_by('-total')
    
    # Top produits vendus
    top_produits = LigneTransaction.objects.filter(
        transaction__date_transaction__date=aujourd_hui,
        transaction__statut='VALIDEE'
    ).values('produit__nom').annotate(
        quantite_totale=Sum('quantite'),
        ca=Sum(F('quantite') * F('prix_unitaire'))
    ).order_by('-quantite_totale')[:10]
    
    # CA par caissier
    ca_par_caissier = ventes.values(
        'caissier__first_name', 'caissier__last_name', 'caissier__username'
    ).annotate(
        nb_ventes=Count('id'),
        ca=Sum('montant_final')
    ).order_by('-ca')
    
    context = {
        'date': aujourd_hui,
        'nb_transactions': nb_transactions,
        'ca_total': ca_total,
        'remises_total': remises_total,
        'panier_moyen': panier_moyen,
        'moyens_paiement': moyens_paiement,
        'top_produits': top_produits,
        'ca_par_caissier': ca_par_caissier,
    }
    
    return render(request, 'caisse/rapport_journalier.html', context)


@login_required
def caissier_mes_ventes(request):
    """Historique des ventes du caissier connecté"""
    if request.user.role not in ['CAISSIER', 'MANAGER', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    from datetime import timedelta
    import calendar
    
    # Paramètres de filtrage
    date_debut_param = request.GET.get('date_debut', '')
    date_fin_param = request.GET.get('date_fin', '')
    periode = request.GET.get('periode', 'aujourd_hui')
    
    # Définir les dates selon la période
    now = timezone.now()
    
    if date_debut_param and date_fin_param:
        date_debut = datetime.strptime(date_debut_param, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_param, '%Y-%m-%d').date()
    elif periode == 'aujourd_hui':
        date_debut = now.date()
        date_fin = now.date()
    elif periode == 'semaine':
        date_debut = now.date() - timedelta(days=now.weekday())
        date_fin = date_debut + timedelta(days=6)
    elif periode == 'mois':
        date_debut = now.date().replace(day=1)
        _, last_day = calendar.monthrange(now.year, now.month)
        date_fin = now.date().replace(day=last_day)
    else:
        date_debut = now.date()
        date_fin = now.date()
    
    # Récupérer les ventes du caissier (utiliser Transaction ou Vente selon votre modèle)
    # Si vous avez Transaction:
    try:
        ventes = Transaction.objects.filter(
            caissier=request.user,
            date_transaction__date__gte=date_debut,
            date_transaction__date__lte=date_fin,
            statut='VALIDEE'
        ).select_related('client').prefetch_related('lignes__produit', 'paiements__type_paiement').order_by('-date_transaction')
        
        # Statistiques
        nb_transactions = ventes.count()
        ca_total = ventes.aggregate(total=Sum('montant_final'))['total'] or 0
        remises_total = ventes.aggregate(total=Sum('montant_remise'))['total'] or 0  # ✅ Corrigé
        panier_moyen = ca_total / nb_transactions if nb_transactions > 0 else 0
        
        # Répartition moyens de paiement
        moyens_paiement = ventes.values('paiements__type_paiement__nom').annotate(
            total=Sum('paiements__montant'),
            nb=Count('id', distinct=True)
        ).order_by('-total')
        
        # Top 10 produits vendus
        from django.db.models import F
        top_produits = LigneTransaction.objects.filter(
            transaction__caissier=request.user,
            transaction__date_transaction__date__gte=date_debut,
            transaction__date_transaction__date__lte=date_fin,
            transaction__statut='VALIDEE'
        ).values('produit__nom', 'produit__reference').annotate(
            quantite_totale=Sum('quantite'),
            ca=Sum(F('quantite') * F('prix_unitaire'))
        ).order_by('-quantite_totale')[:10]
        
        # CA par jour (pour graphique)
        ca_par_jour = ventes.extra(
            select={'jour': 'DATE(date_transaction)'}
        ).values('jour').annotate(
            ca=Sum('montant_final'),
            nb_ventes=Count('id')
        ).order_by('jour')
        
    except NameError:
        # Si Transaction n'existe pas, utiliser Vente
        ventes = Vente.objects.filter(
            caissier=request.user,
            date_vente__date__gte=date_debut,
            date_vente__date__lte=date_fin
        ).select_related('client').prefetch_related('lignes__produit').order_by('-date_vente')
        
        # Statistiques
        nb_transactions = ventes.count()
        ca_total = ventes.aggregate(total=Sum('montant_final'))['total'] or 0
        remises_total = ventes.aggregate(total=Sum('remise'))['total'] or 0
        panier_moyen = ca_total / nb_transactions if nb_transactions > 0 else 0
        
        # Répartition moyens de paiement
        moyens_paiement = ventes.values('moyen_paiement').annotate(
            total=Sum('montant_final'),
            nb=Count('id')
        ).order_by('-total')
        
        # Top 10 produits vendus
        from django.db.models import F
        top_produits = LigneVente.objects.filter(
            vente__caissier=request.user,
            vente__date_vente__date__gte=date_debut,
            vente__date_vente__date__lte=date_fin
        ).values('produit__nom', 'produit__reference').annotate(
            quantite_totale=Sum('quantite'),
            ca=Sum(F('quantite') * F('prix_unitaire'))
        ).order_by('-quantite_totale')[:10]
        
        # CA par jour (pour graphique)
        ca_par_jour = ventes.extra(
            select={'jour': 'DATE(date_vente)'}
        ).values('jour').annotate(
            ca=Sum('montant_final'),
            nb_ventes=Count('id')
        ).order_by('jour')
    
    context = {
        'transactions': ventes,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'periode': periode,
        'nb_transactions': nb_transactions,
        'ca_total': ca_total,
        'remises_total': remises_total,
        'panier_moyen': panier_moyen,
        'moyens_paiement': moyens_paiement,
        'top_produits': top_produits,
        'ca_par_jour': list(ca_par_jour),
    }
    
    return render(request, 'caisse/mes_ventes.html', context)


# ==================== GESTION PLANNING & CONGÉS (Scénario 8.1.3) ====================

@login_required
def mon_planning(request):
    """Espace employé : Consulter son planning"""
    employe = request.user
    
    # Planning de la semaine en cours
    aujourd_hui = timezone.now().date()
    debut_semaine = aujourd_hui - timedelta(days=aujourd_hui.weekday())
    fin_semaine = debut_semaine + timedelta(days=6)
    
    plannings = Planning.objects.filter(
        employe=employe,
        date__gte=debut_semaine,
        date__lte=fin_semaine
    ).order_by('date', 'creneau')
    
    # Mes demandes de congés en attente
    demandes_attente = DemandeConge.objects.filter(
        employe=employe,
        statut='EN_ATTENTE'
    ).order_by('-cree_le')[:5]
    
    # Mes absences enregistrées
    absences = Absence.objects.filter(
        employe=employe
    ).order_by('-date')[:5]
    
    # Heures travaillées ce mois
    debut_mois = aujourd_hui.replace(day=1)
    heures_mois = Planning.objects.filter(
        employe=employe,
        date__gte=debut_mois,
        date__lte=aujourd_hui
    ).aggregate(
        total_heures=Sum('heures_travaillees')
    )['total_heures'] or 0
    
    context = {
        'employe': employe,
        'plannings': plannings,
        'demandes_attente': demandes_attente,
        'absences': absences,
        'heures_mois': heures_mois,
        'debut_semaine': debut_semaine,
        'fin_semaine': fin_semaine,
    }
    
    return render(request, 'planning/mon_planning.html', context)


@login_required
def demander_conge(request):
    """Espace employé : Faire une demande de congé"""
    employe = request.user
    
    if request.method == 'POST':
        type_conge = request.POST.get('type_conge')
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        motif = request.POST.get('motif', '')
        
        try:
            # Convertir les dates string en objets date
            from datetime import datetime
            date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            
            # Créer la demande
            demande = DemandeConge.objects.create(
                employe=employe,
                type_conge=type_conge,
                date_debut=date_debut,
                date_fin=date_fin,
                motif=motif,
                statut='EN_ATTENTE'
            )
            
            messages.success(request, f'✅ Demande de congé envoyée ! Référence: {demande.id}')
            return redirect('mes_demandes_conges')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la demande: {str(e)}')
    
    context = {
        'employe': employe,
        'types_conges': DemandeConge.TYPES_CONGE,
    }
    
    return render(request, 'planning/demander_conge.html', context)


@login_required
def mes_demandes_conges(request):
    """Espace employé : Historique de mes demandes"""
    employe = request.user
    
    demandes = DemandeConge.objects.filter(
        employe=employe
    ).order_by('-cree_le')
    
    context = {
        'demandes': demandes,
        'employe': employe,
    }
    
    return render(request, 'planning/mes_demandes.html', context)


@login_required
def changer_mot_de_passe(request):
    """Espace employé : Modifier son mot de passe"""
    if request.method == 'POST':
        ancien_mdp = request.POST.get('ancien_mot_de_passe')
        nouveau_mdp = request.POST.get('nouveau_mot_de_passe')
        confirmer_mdp = request.POST.get('confirmer_mot_de_passe')
        
        # Vérifier l'ancien mot de passe
        if not request.user.check_password(ancien_mdp):
            messages.error(request, '❌ Ancien mot de passe incorrect.')
            return redirect('changer_mot_de_passe')
        
        # Vérifier que les nouveaux mots de passe correspondent
        if nouveau_mdp != confirmer_mdp:
            messages.error(request, '❌ Les nouveaux mots de passe ne correspondent pas.')
            return redirect('changer_mot_de_passe')
        
        # Vérifier la longueur minimale
        if len(nouveau_mdp) < 6:
            messages.error(request, '❌ Le mot de passe doit contenir au moins 6 caractères.')
            return redirect('changer_mot_de_passe')
        
        # Changer le mot de passe
        request.user.set_password(nouveau_mdp)
        request.user.save()
        
        # Re-authentifier l'utilisateur
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)
        
        messages.success(request, '✅ Mot de passe modifié avec succès !')
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    return render(request, 'planning/changer_mot_de_passe.html')


# ==================== ESPACE RH/MANAGER ====================

@login_required
def rh_demandes_conges(request):
    """RH/Manager : Gérer les demandes de congés"""
    if request.user.role not in ['RH', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "Accès refusé. Réservé aux RH/Managers.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    # Filtrer par statut
    statut_filtre = request.GET.get('statut', 'EN_ATTENTE')
    
    demandes = DemandeConge.objects.filter(
        statut=statut_filtre
    ).select_related('employe').order_by('-cree_le')
    
    # Statistiques
    nb_en_attente = DemandeConge.objects.filter(statut='EN_ATTENTE').count()
    nb_approuvees = DemandeConge.objects.filter(statut='APPROUVE').count()
    nb_rejetees = DemandeConge.objects.filter(statut='REFUSE').count()  # Corrigé: REFUSE pas REJETE
    
    context = {
        'demandes': demandes,
        'statut_filtre': statut_filtre,
        'nb_en_attente': nb_en_attente,
        'nb_approuvees': nb_approuvees,
        'nb_rejetees': nb_rejetees,
    }
    
    return render(request, 'planning/rh_demandes_conges.html', context)


@login_required
def rh_traiter_demande(request, demande_id):
    """RH/Manager : Approuver ou rejeter une demande"""
    if request.user.role not in ['RH', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    try:
        demande = DemandeConge.objects.get(id=demande_id)
        
        if request.method == 'POST':
            action = request.POST.get('action')
            commentaire_rh = request.POST.get('commentaire_rh', '')
            
            if action == 'approuver':
                demande.statut = 'APPROUVE'
                demande.approuve_par = request.user.employe if hasattr(request.user, 'employe') else None
                demande.date_traitement = timezone.now()
                demande.commentaire_rh = commentaire_rh
                demande.save()
                
                messages.success(request, f'✅ Demande de congé approuvée pour {demande.employe.nom}')
                
            elif action == 'rejeter':
                demande.statut = 'REJETE'
                demande.approuve_par = request.user.employe if hasattr(request.user, 'employe') else None
                demande.date_traitement = timezone.now()
                demande.commentaire_rh = commentaire_rh
                demande.save()
                
                messages.warning(request, f'❌ Demande de congé rejetée pour {demande.employe.nom}')
            
            return redirect('rh_demandes_conges')
            
    except DemandeConge.DoesNotExist:
        messages.error(request, 'Demande introuvable.')
    
    return redirect('rh_demandes_conges')


@login_required
def rh_gestion_absences(request):
    """RH/Manager : Gérer les absences"""
    if request.user.role not in ['RH', 'MANAGER', 'ADMIN']:
        messages.error(request, "Accès refusé. Réservé aux RH/Managers.")
        return redirect('dashboard')
    
    if request.method == 'POST':
        employe_id = request.POST.get('employe_id')
        date_absence = request.POST.get('date')
        type_absence = request.POST.get('type_absence')
        justifiee = request.POST.get('justifiee') == 'on'
        commentaire = request.POST.get('commentaire', '')
        
        try:
            employe = Employe.objects.get(id=employe_id)
            
            Absence.objects.create(
                employe=employe,
                date=date_absence,
                type_absence=type_absence,
                justifiee=justifiee,
                commentaire=commentaire
            )
            
            messages.success(request, f'✅ Absence enregistrée pour {employe.nom}')
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    # Liste des absences récentes
    absences = Absence.objects.select_related('employe').order_by('-date')[:50]
    
    # Employés actifs
    employes = Employe.objects.filter(est_actif=True).order_by('nom')
    
    context = {
        'absences': absences,
        'employes': employes,
        'types_absences': Absence.TYPE_ABSENCE_CHOICES,
    }
    
    return render(request, 'planning/rh_gestion_absences.html', context)


@login_required
def rh_reinitialiser_mdp(request):
    """RH/Manager : Réinitialiser le mot de passe d'un employé"""
    if request.user.role not in ['RH', 'MANAGER', 'ADMIN', 'DG', 'DAF']:
        messages.error(request, "Accès refusé. Réservé aux RH/Managers.")
        return redirect('dashboard_rh')
    
    if request.method == 'POST':
        employe_id = request.POST.get('employe_id')
        nouveau_mdp = request.POST.get('nouveau_mot_de_passe')
        
        try:
            employe = Employe.objects.get(id=employe_id)
            
            # Réinitialiser le mot de passe
            employe.set_password(nouveau_mdp)
            employe.save()
            
            messages.success(request, f'✅ Mot de passe réinitialisé pour {employe.get_full_name()} ({employe.employee_id})')
            messages.info(request, f'Nouveau mot de passe: {nouveau_mdp}')
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    # Liste des employés
    employes = Employe.objects.filter(est_actif=True).order_by('first_name', 'last_name')
    
    context = {
        'employes': employes,
    }
    
    return render(request, 'planning/rh_reinitialiser_mdp.html', context)


# ==================== GESTION DES COUPONS (Marketing) ====================

@login_required
def marketing_coupons_list(request):
    """Liste des coupons (Marketing)"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    # Filtres
    type_coupon = request.GET.get('type', 'all')
    statut = request.GET.get('statut', 'all')
    
    coupons = Coupon.objects.all()
    
    if type_coupon != 'all':
        coupons = coupons.filter(type_coupon=type_coupon)
    
    if statut != 'all':
        coupons = coupons.filter(statut=statut)
    
    # Statistiques
    stats = {
        'total': Coupon.objects.count(),
        'actifs': Coupon.objects.filter(statut='ACTIF').count(),
        'generiques': Coupon.objects.filter(type_coupon='GENERIC').count(),
        'speciaux': Coupon.objects.filter(type_coupon='SPECIAL').count(),
        'utilisations': UtilisationCoupon.objects.count(),
        'remise_totale': UtilisationCoupon.objects.aggregate(total=Sum('montant_remise'))['total'] or 0
    }
    
    context = {
        'coupons': coupons,
        'stats': stats,
        'type_filter': type_coupon,
        'statut_filter': statut,
    }
    
    return render(request, 'marketing/coupons_list.html', context)


@login_required
def marketing_coupon_create(request):
    """Créer un nouveau coupon"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    if request.method == 'POST':
        type_coupon = request.POST.get('type_coupon')
        type_remise = request.POST.get('type_remise')
        valeur = request.POST.get('valeur')
        description = request.POST.get('description')
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        montant_minimum = request.POST.get('montant_minimum', 0)
        niveau_requis = request.POST.get('niveau_fidelite_requis', None)
        limite_globale = request.POST.get('limite_globale', None)
        
        try:
            coupon = Coupon.objects.create(
                type_coupon=type_coupon,
                type_remise=type_remise,
                valeur=valeur,
                description=description,
                date_debut=datetime.strptime(date_debut, '%Y-%m-%d').date(),
                date_fin=datetime.strptime(date_fin, '%Y-%m-%d').date(),
                montant_minimum=montant_minimum or 0,
                niveau_fidelite_requis=niveau_requis if niveau_requis else None,
                limite_globale=int(limite_globale) if limite_globale else None,
                cree_par=request.user,
                statut='ACTIF'
            )
            
            messages.success(request, f'Coupon {coupon.code} créé avec succès!')
            return redirect('marketing_coupons_list')
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    context = {
        'niveaux_fidelite': Client.NIVEAUX_FIDELITE,
    }
    
    return render(request, 'marketing/coupon_create.html', context)


@login_required
def marketing_coupon_generer_speciaux(request):
    """Générer des coupons spéciaux pour les clients fidèles"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    if request.method == 'POST':
        from datetime import date, timedelta
        
        niveau_min = request.POST.get('niveau_minimum', 'SILVER')
        type_remise = request.POST.get('type_remise', 'POURCENTAGE')
        valeur = float(request.POST.get('valeur', 0))
        description = request.POST.get('description')
        duree_jours = int(request.POST.get('duree_jours', 30))
        
        # Sélectionner les clients éligibles
        niveaux_eligibles = {
            'SILVER': ['SILVER', 'GOLD', 'VIP'],
            'GOLD': ['GOLD', 'VIP'],
            'VIP': ['VIP']
        }
        
        clients = Client.objects.filter(
            niveau_fidelite__in=niveaux_eligibles.get(niveau_min, ['VIP']),
            est_actif=True
        )
        
        nb_coupons = 0
        date_debut = date.today()
        date_fin = date_debut + timedelta(days=duree_jours)
        
        for client in clients:
            # Vérifier si le client n'a pas déjà un coupon similaire actif
            coupon_existant = Coupon.objects.filter(
                client=client,
                type_coupon='SPECIAL',
                statut='ACTIF',
                est_utilise=False,
                date_fin__gte=date.today()
            ).exists()
            
            if not coupon_existant:
                Coupon.objects.create(
                    type_coupon='SPECIAL',
                    type_remise=type_remise,
                    valeur=valeur,
                    description=description,
                    date_debut=date_debut,
                    date_fin=date_fin,
                    client=client,
                    cree_par=request.user,
                    statut='ACTIF'
                )
                nb_coupons += 1
        
        messages.success(request, f'{nb_coupons} coupons spéciaux générés avec succès!')
        return redirect('marketing_coupons_list')
    
    context = {
        'niveaux': ['SILVER', 'GOLD', 'VIP']
    }
    
    return render(request, 'marketing/coupon_generer_speciaux.html', context)


@login_required
def marketing_coupon_desactiver(request, coupon_id):
    """Désactiver un coupon"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        return JsonResponse({'success': False, 'error': 'Accès refusé'})
    
    coupon = get_object_or_404(Coupon, id=coupon_id)
    coupon.statut = 'DESACTIVE'
    coupon.save()
    
    messages.success(request, f'Coupon {coupon.code} désactivé.')
    return redirect('marketing_coupons_list')


@login_required
def caisse_valider_coupon(request):
    """AJAX - Valider un coupon au POS"""
    if request.method == 'POST':
        code_coupon = request.POST.get('code', '').strip().upper()
        client_id = request.POST.get('client_id')
        montant_achat = float(request.POST.get('montant', 0))
        
        try:
            coupon = Coupon.objects.get(code=code_coupon)
            
            # Récupérer le client si fourni
            client = None
            if client_id:
                client = Client.objects.get(id=client_id)
            
            # Valider le coupon
            valide, message = coupon.est_valide(client, montant_achat)
            
            if not valide:
                return JsonResponse({
                    'success': False,
                    'error': message
                })
            
            # Calculer la remise
            remise = coupon.calculer_remise(montant_achat)
            
            return JsonResponse({
                'success': True,
                'coupon': {
                    'id': coupon.id,
                    'code': coupon.code,
                    'description': coupon.description,
                    'type_remise': coupon.type_remise,
                    'valeur': float(coupon.valeur),
                    'remise': float(remise)
                }
            })
            
        except Coupon.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Coupon invalide'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


# ==================== ALGORITHME INTELLIGENT DE FIDÉLITÉ ====================

@login_required
def marketing_analyser_fidelite(request):
    """Lancer l'analyse intelligente de fidélité"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    if request.method == 'POST':
        from django.core.management import call_command
        from io import StringIO
        
        generer_coupons = request.POST.get('generer_coupons') == 'on'
        
        # Capturer la sortie de la commande
        out = StringIO()
        
        try:
            call_command('analyser_fidelite', 
                        generer_coupons=generer_coupons,
                        stdout=out)
            
            # Récupérer les statistiques
            output = out.getvalue()
            
            messages.success(request, '✅ Analyse de fidélité terminée avec succès!')
            
            context = {
                'output': output,
                'success': True
            }
            
            return render(request, 'marketing/analyse_fidelite_resultat.html', context)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'analyse: {str(e)}')
            return redirect('marketing_analyser_fidelite')
    
    # Afficher la page de confirmation
    # Statistiques avant analyse
    from django.db.models import Count
    
    stats_avant = {
        'tous': Client.objects.filter(niveau_fidelite='TOUS', est_actif=True).count(),
        'silver': Client.objects.filter(niveau_fidelite='SILVER', est_actif=True).count(),
        'gold': Client.objects.filter(niveau_fidelite='GOLD', est_actif=True).count(),
        'vip': Client.objects.filter(niveau_fidelite='VIP', est_actif=True).count(),
    }
    stats_avant['total'] = sum(stats_avant.values())
    
    context = {
        'stats_avant': stats_avant
    }
    
    return render(request, 'marketing/analyse_fidelite.html', context)


@login_required
def marketing_fidelite_stats(request):
    """Statistiques détaillées sur la fidélité"""
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN']:
        messages.error(request, "Accès refusé.")
        dashboard_url = get_dashboard_by_role(request.user)
        return redirect(dashboard_url)
    
    from datetime import timedelta
    
    today = timezone.now().date()
    six_mois_ago = today - timedelta(days=180)
    trois_mois_ago = today - timedelta(days=90)
    un_mois_ago = today - timedelta(days=30)
    
    # Répartition par niveau
    repartition = {
        'TOUS': Client.objects.filter(niveau_fidelite='TOUS', est_actif=True).count(),
        'SILVER': Client.objects.filter(niveau_fidelite='SILVER', est_actif=True).count(),
        'GOLD': Client.objects.filter(niveau_fidelite='GOLD', est_actif=True).count(),
        'VIP': Client.objects.filter(niveau_fidelite='VIP', est_actif=True).count(),
    }
    
    # Essayer Transaction d'abord
    try:
        # Clients actifs (achat dans les 30 derniers jours)
        clients_actifs = Client.objects.filter(
            transactions__statut='VALIDEE',
            transactions__date_transaction__date__gte=un_mois_ago,
            est_actif=True
        ).distinct().count()
        
        # CA par niveau sur 3 mois
        ca_par_niveau = {}
        for niveau in ['TOUS', 'SILVER', 'GOLD', 'VIP']:
            ca = Transaction.objects.filter(
                client__niveau_fidelite=niveau,
                statut='VALIDEE',
                date_transaction__date__gte=trois_mois_ago
            ).aggregate(total=Sum('montant_final'))['total'] or 0
            
            nb_clients = repartition[niveau]
            ca_par_niveau[niveau] = {
                'ca_total': ca,
                'ca_moyen': ca / nb_clients if nb_clients > 0 else 0
            }
        
        # Top 10 clients VIP
        top_vip = Transaction.objects.filter(
            client__niveau_fidelite='VIP',
            statut='VALIDEE',
            date_transaction__date__gte=trois_mois_ago
        ).values(
            'client__nom', 'client__prenom', 'client__telephone'
        ).annotate(
            ca=Sum('montant_final'),
            nb_achats=Count('id')
        ).order_by('-ca')[:10]
        
    except:
        # Si Transaction n'existe pas, utiliser Vente
        clients_actifs = Client.objects.filter(
            ventes__date_vente__date__gte=un_mois_ago,
            est_actif=True
        ).distinct().count()
        
        ca_par_niveau = {}
        for niveau in ['TOUS', 'SILVER', 'GOLD', 'VIP']:
            ca = Vente.objects.filter(
                client__niveau_fidelite=niveau,
                date_vente__date__gte=trois_mois_ago
            ).aggregate(total=Sum('montant_final'))['total'] or 0
            
            nb_clients = repartition[niveau]
            ca_par_niveau[niveau] = {
                'ca_total': ca,
                'ca_moyen': ca / nb_clients if nb_clients > 0 else 0
            }
        
        top_vip = Vente.objects.filter(
            client__niveau_fidelite='VIP',
            date_vente__date__gte=trois_mois_ago
        ).values(
            'client__nom', 'client__prenom', 'client__telephone'
        ).annotate(
            ca=Sum('montant_final'),
            nb_achats=Count('id')
        ).order_by('-ca')[:10]
    
    context = {
        'repartition': repartition,
        'clients_actifs': clients_actifs,
        'ca_par_niveau': ca_par_niveau,
        'top_vip': top_vip,
    }
    
    return render(request, 'marketing/fidelite_stats.html', context)


@login_required
def marketing_dashboard_kpis(request):
    """
    Dashboard consolidé avec tous les KPIs CRM et fidélité
    """
    # Vérifier les autorisations
    if not hasattr(request.user, 'role'):
        return JsonResponse({'success': False, 'error': 'Utilisateur non autorisé'}, status=403)
    
    if request.user.role not in ['MARKETING', 'DG', 'ADMIN', 'MANAGER', 'DAF']:
        return JsonResponse({'success': False, 'error': 'Accès refusé'}, status=403)
    
    # Récupérer les filtres
    periode = request.GET.get('periode', '30')  # Par défaut 30 jours
    segment = request.GET.get('segment', 'tous')  # tous, VIP, GOLD, SILVER, TOUS
    
    try:
        jours = int(periode)
    except:
        jours = 30
    
    date_debut = timezone.now() - timedelta(days=jours)
    
    # ==================== KPI 1: TAUX D'IDENTIFICATION ====================
    try:
        # Essayer avec Transaction d'abord
        total_transactions = Transaction.objects.filter(
            date_transaction__gte=date_debut
        ).count()
        
        transactions_identifiees = Transaction.objects.filter(
            date_transaction__gte=date_debut,
            client__isnull=False
        ).count()
    except:
        # Fallback sur Vente
        total_transactions = Vente.objects.filter(
            date_vente__gte=date_debut
        ).count()
        
        transactions_identifiees = Vente.objects.filter(
            date_vente__gte=date_debut
        ).exclude(
            Q(telephone_client__isnull=True) | Q(telephone_client='')
        ).count()
    
    taux_identification = (transactions_identifiees / total_transactions * 100) if total_transactions > 0 else 0
    
    # Évolution sur 7 périodes (pour graphique)
    evolution_identification = []
    for i in range(6, -1, -1):
        debut_periode = timezone.now() - timedelta(days=jours * (i + 1))
        fin_periode = timezone.now() - timedelta(days=jours * i)
        
        try:
            total_p = Transaction.objects.filter(
                date_transaction__gte=debut_periode,
                date_transaction__lt=fin_periode
            ).count()
            
            ident_p = Transaction.objects.filter(
                date_transaction__gte=debut_periode,
                date_transaction__lt=fin_periode,
                client__isnull=False
            ).count()
        except:
            total_p = Vente.objects.filter(
                date_vente__gte=debut_periode,
                date_vente__lt=fin_periode
            ).count()
            
            ident_p = Vente.objects.filter(
                date_vente__gte=debut_periode,
                date_vente__lt=fin_periode
            ).exclude(
                Q(telephone_client__isnull=True) | Q(telephone_client='')
            ).count()
        
        taux_p = (ident_p / total_p * 100) if total_p > 0 else 0
        evolution_identification.append({
            'periode': f"P{7-i}",
            'taux': round(taux_p, 1)
        })
    
    # ==================== KPI 2: CROISSANCE SEGMENTS ====================
    clients_par_niveau = Client.objects.filter(
        telephone__isnull=False
    ).values('niveau_fidelite').annotate(
        count=Count('id')
    )
    
    repartition_actuelle = {
        'VIP': 0,
        'GOLD': 0,
        'SILVER': 0,
        'TOUS': 0
    }
    
    for item in clients_par_niveau:
        niveau = item['niveau_fidelite'] or 'TOUS'
        repartition_actuelle[niveau] = item['count']
    
    # Évolution des segments (6 derniers mois)
    evolution_segments = []
    for i in range(5, -1, -1):
        mois = timezone.now() - timedelta(days=30 * i)
        
        # Compter les clients ayant fait au moins 1 achat jusqu'à ce mois
        try:
            clients_actifs = Transaction.objects.filter(
                date_transaction__lte=mois,
                client__isnull=False
            ).values('client__niveau_fidelite').annotate(
                count=Count('client', distinct=True)
            )
        except:
            # Approximation avec Vente
            clients_actifs = []
        
        repartition_mois = {
            'VIP': 0,
            'GOLD': 0,
            'SILVER': 0,
            'TOUS': 0
        }
        
        for item in clients_actifs:
            niveau = item['client__niveau_fidelite'] or 'TOUS'
            if niveau in repartition_mois:
                repartition_mois[niveau] = item['count']
        
        evolution_segments.append({
            'mois': mois.strftime('%b'),
            'vip': repartition_mois['VIP'],
            'gold': repartition_mois['GOLD'],
            'silver': repartition_mois['SILVER'],
            'tous': repartition_mois['TOUS']
        })
    
    # ==================== KPI 3: TAUX UTILISATION COUPONS ====================
    coupons_distribues = Coupon.objects.filter(
        date_creation__gte=date_debut,
        statut='ACTIF'
    ).count()
    
    coupons_utilises = UtilisationCoupon.objects.filter(
        date_utilisation__gte=date_debut
    ).values('coupon').distinct().count()
    
    taux_utilisation_coupons = (coupons_utilises / coupons_distribues * 100) if coupons_distribues > 0 else 0
    
    # Par type
    coupons_generiques = Coupon.objects.filter(
        date_creation__gte=date_debut,
        type_coupon='GENERIC',
        statut='ACTIF'
    ).count()
    
    coupons_speciaux = Coupon.objects.filter(
        date_creation__gte=date_debut,
        type_coupon='SPECIAL',
        statut='ACTIF'
    ).count()
    
    utilisations_generiques = UtilisationCoupon.objects.filter(
        date_utilisation__gte=date_debut,
        coupon__type_coupon='GENERIC'
    ).values('coupon').distinct().count()
    
    utilisations_speciaux = UtilisationCoupon.objects.filter(
        date_utilisation__gte=date_debut,
        coupon__type_coupon='SPECIAL'
    ).values('coupon').distinct().count()
    
    taux_generiques = (utilisations_generiques / coupons_generiques * 100) if coupons_generiques > 0 else 0
    taux_speciaux = (utilisations_speciaux / coupons_speciaux * 100) if coupons_speciaux > 0 else 0
    
    # ==================== KPI 4: MARGE NETTE APRÈS REMISES ====================
    try:
        # CA Brut
        ca_brut = Transaction.objects.filter(
            date_transaction__gte=date_debut
        ).aggregate(
            total=Sum('montant_total')
        )['total'] or 0
        
        # Remises fidélité
        remises_fidelite = Transaction.objects.filter(
            date_transaction__gte=date_debut,
            client__isnull=False
        ).aggregate(
            total=Sum('remise_fidelite')
        )['total'] or 0
        
        # Remises coupons
        remises_coupons = UtilisationCoupon.objects.filter(
            date_utilisation__gte=date_debut
        ).aggregate(
            total=Sum('montant_remise')
        )['total'] or 0
    except:
        ca_brut = Vente.objects.filter(
            date_vente__gte=date_debut
        ).aggregate(
            total=Sum('montant_total')
        )['total'] or 0
        
        remises_fidelite = 0
        remises_coupons = 0
    
    total_remises = remises_fidelite + remises_coupons
    ca_net = ca_brut - total_remises
    taux_remise = (total_remises / ca_brut * 100) if ca_brut > 0 else 0
    
    # ==================== KPI 5: RÉTENTION CLIENTS ====================
    # Clients actifs mois actuel
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    try:
        clients_mois_actuel = Transaction.objects.filter(
            date_transaction__gte=debut_mois,
            client__isnull=False
        ).values('client').distinct().count()
        
        # Clients actifs mois précédent
        debut_mois_precedent = (debut_mois - timedelta(days=1)).replace(day=1)
        fin_mois_precedent = debut_mois - timedelta(seconds=1)
        
        clients_mois_precedent = Transaction.objects.filter(
            date_transaction__gte=debut_mois_precedent,
            date_transaction__lte=fin_mois_precedent,
            client__isnull=False
        ).values('client').distinct().count()
    except:
        clients_mois_actuel = 0
        clients_mois_precedent = 0
    
    taux_retention = (clients_mois_actuel / clients_mois_precedent * 100) if clients_mois_precedent > 0 else 0
    taux_churn = 100 - taux_retention
    
    # Évolution rétention
    evolution_retention = []
    for i in range(5, -1, -1):
        mois = timezone.now() - timedelta(days=30 * i)
        debut_m = mois.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        fin_m = (debut_m + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
        
        try:
            actifs_m = Transaction.objects.filter(
                date_transaction__gte=debut_m,
                date_transaction__lte=fin_m,
                client__isnull=False
            ).values('client').distinct().count()
        except:
            actifs_m = 0
        
        evolution_retention.append({
            'mois': debut_m.strftime('%b'),
            'clients': actifs_m
        })
    
    # ==================== KPI 6: FRÉQUENCE D'ACHAT ====================
    # Par niveau de fidélité
    frequences = {}
    
    for niveau in ['VIP', 'GOLD', 'SILVER', 'TOUS']:
        clients_niveau = Client.objects.filter(niveau_fidelite=niveau)
        
        try:
            nb_achats = Transaction.objects.filter(
                date_transaction__gte=date_debut,
                client__niveau_fidelite=niveau
            ).count()
            
            nb_clients = Transaction.objects.filter(
                date_transaction__gte=date_debut,
                client__niveau_fidelite=niveau
            ).values('client').distinct().count()
        except:
            nb_achats = 0
            nb_clients = 0
        
        freq = (nb_achats / nb_clients) if nb_clients > 0 else 0
        frequences[niveau] = round(freq / (jours / 30), 2)  # Achats par mois
    
    # ==================== TOP COUPONS ====================
    top_coupons = UtilisationCoupon.objects.filter(
        date_utilisation__gte=date_debut
    ).values(
        'coupon__code',
        'coupon__type_remise',
        'coupon__valeur_remise'
    ).annotate(
        nb_utilisations=Count('id'),
        total_remise=Sum('montant_remise')
    ).order_by('-nb_utilisations')[:10]
    
    # ==================== CA PAR SEGMENT (filtré) ====================
    ca_par_segment = {}
    
    if segment == 'tous':
        niveaux = ['VIP', 'GOLD', 'SILVER', 'TOUS']
    else:
        niveaux = [segment]
    
    for niveau in niveaux:
        try:
            ca = Transaction.objects.filter(
                date_transaction__gte=date_debut,
                client__niveau_fidelite=niveau
            ).aggregate(total=Sum('montant_total'))['total'] or 0
        except:
            ca = 0
        
        ca_par_segment[niveau] = ca
    
    context = {
        'periode': jours,
        'segment': segment,
        'date_debut': date_debut,
        
        # KPI 1
        'taux_identification': round(taux_identification, 1),
        'transactions_identifiees': transactions_identifiees,
        'total_transactions': total_transactions,
        'evolution_identification': evolution_identification,
        
        # KPI 2
        'repartition_actuelle': repartition_actuelle,
        'evolution_segments': evolution_segments,
        'total_clients': sum(repartition_actuelle.values()),
        
        # KPI 3
        'taux_utilisation_coupons': round(taux_utilisation_coupons, 1),
        'coupons_distribues': coupons_distribues,
        'coupons_utilises': coupons_utilises,
        'taux_generiques': round(taux_generiques, 1),
        'taux_speciaux': round(taux_speciaux, 1),
        'coupons_generiques': coupons_generiques,
        'coupons_speciaux': coupons_speciaux,
        
        # KPI 4
        'ca_brut': ca_brut,
        'ca_net': ca_net,
        'total_remises': total_remises,
        'remises_fidelite': remises_fidelite,
        'remises_coupons': remises_coupons,
        'taux_remise': round(taux_remise, 1),
        
        # KPI 5
        'taux_retention': round(taux_retention, 1),
        'taux_churn': round(taux_churn, 1),
        'clients_mois_actuel': clients_mois_actuel,
        'clients_mois_precedent': clients_mois_precedent,
        'evolution_retention': evolution_retention,
        
        # KPI 6
        'frequences': frequences,
        
        # Extras
        'top_coupons': top_coupons,
        'ca_par_segment': ca_par_segment,
    }
    
    return render(request, 'marketing/dashboard_kpis.html', context)
